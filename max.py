#!/usr/bin/env python3
"""
Max - wide procurement dataset builder
======================================

Author  : Prof. Shahab Anbarjafari
Purpose : Assemble one wide, analysis-ready table from the separate procurement
          extracts, in six stages, each written to CSV and JSONL.

    Stage 1   Sievo transactions widened with the matching invoice lines.
    Stage 2   The result widened again with the Maximo or Basware purchase
              order line, matched on PO number plus PO line number.
    Stage 3   The free text carried on every row read and turned into
              structured columns.
    Stage 4   Agent 1's standardised English description, language reading and
              material-or-service call added to every row.
    Stage 5   Agent 2's purchase group added, so every line carries the
              category-L5 grouping it belongs to.
    Stage 6   Agent 3's catalogue match and standard-item status added. This is
              the complete table, and the one to hand over.

Agent 4 does not appear as a stage. Its unit of analysis is a supplier within a
comparison scope rather than a purchase line, so it is written alongside as
`max_supplier_consolidation`, joinable on the `Supplier_Key` that stage 6 adds
to every row. Folding it into the wide table would mean either multiplying the
rows or inventing a per-line summary of a supplier-level finding.

Each stage is a file, so a reviewer can see exactly what each step contributed
and, if one is wrong, fall back to the one before it. Where Max's own stage 3
and Agent 1 both name a column, the agent's reading wins because it is the one
the client reviewed - but only where it is populated, so that an agent which
failed to resolve a field cannot blank a value Max had already established.
Max's own reading of every such column stays visible in the stage-3 file.

The governing rule
------------------
A join here *widens* the transaction table; it never lengthens it. The Sievo
extract is the spend record, so one Sievo row must remain one output row from
the first stage to the last. Where several invoice lines or several PO lines
answer to the same transaction, they are folded into summary columns and the
fan-out is recorded, rather than being emitted as extra rows. Any other choice
would multiply spend during a join, and a spend figure that changes because a
reference table was attached to it is worse than no figure at all. The row count
is asserted at the end of every stage.

A second rule follows from the first: on a header-level match, a column is
filled only when every candidate line agrees on its value. If two PO lines on
the same order name different suppliers, the supplier column is left empty and
the disagreement is reported, because a plausible-looking wrong value costs more
to find later than a blank does now.

What matches what
-----------------
Nothing is assumed about which file is which. Every table found under the source
folder is classified by its header signature, so the files can be renamed, split
or supplied several at a time. Each join then tries a ladder of key strategies
from the most specific to the least, and records which one fired for each row:

    Stage 1   invoice number + line -> document number + line -> document id +
              line -> the same three again at document level
    Stage 2   PO number + PO line -> PO number where the order has exactly one
              line -> PO number at header level

"Document id" is the opaque identifier Sievo carries in `DocumentIdentifier`,
which also appears as the `docid` parameter of `InvoiceLink` and as the leading
element of the invoice `xml_file_name`. It is worth trying because it survives
in extracts where the invoice number column was never populated.

Every stage writes a diagnostic block recording how many rows matched, by which
strategy, and how populated each candidate key was. When a join matches nothing
that block is the answer to why, and it is the thing to send back to whoever
produced the extract.

The agents
----------
Stages 4 to 6 run `agent1.py`, `agent2.py` and `agent3.py` as separate processes
over the stage-3 file, then join what each produced back onto the transaction
row. They are run rather than imported: each is a standalone tool with its own
cache, its own spend guard and its own manifest that the client reads, and
keeping them in their own process means Max can add their columns without
becoming responsible for their internals. Nothing in an agent is changed to suit
Max.

The one place the two meet is the column names, which are listed explicitly and
checked against what each agent actually wrote. A promised column that did not
arrive stops the run, because a wide table quietly missing a column is worse
than a run that failed and said which one was missing.

An agent that fails costs the columns it would have added and is reported as
having failed; the stages already on disk are unaffected and the last good one
is named as the table to use.

Language model
--------------
Stages 1 and 2 never use one. A join is a key operation with a right answer, and
a model can only introduce error into it.

Stage 3 resolves free text with a controlled vocabulary, compound splitting and
a set of extraction rules, all of which run locally at no cost. The model is
offered only for the residue those cannot read, is asked once per distinct
string rather than once per row, is cached on a hash of the request, and runs
under the same spend guard as the other agents: a limit in dollars is agreed up
front, reaching it pauses the run to ask, and declining finishes the work on the
local stack.

Being stopped part way through
-----------------------------
A run that is interrupted keeps what it finished. Each stage is recorded once its
files are written and closed, together with their size and modification time, and
running the same command again carries those stages over and starts at the one
that did not finish. This matters because the work is not evenly spread: stage 3
reads every distinct string in the extract and the agents after it run for hours,
so a build stopped near the end has usually earned nearly all of it.

Only a stage that finished is ever carried over. A part-written table is not
recorded, is deleted when the run resumes, and cannot be mistaken for a result. If
the extracts, the vocabulary or the settings that shape these stages have changed
since, the record no longer applies and the run says which of them changed and
starts from stage 1. `--restart` does that on demand.

Usage
-----
    python max.py                 # prompts for each path in turn
    python max.py --non-interactive --sources ./sources --results ./results
    python max.py --non-interactive --no-agents      # stop after stage 3
    python max.py --non-interactive --restart        # ignore an interrupted run
    python max.py --help
"""

from __future__ import annotations

import argparse
import copy
import csv
import datetime
import hashlib
import io
import json
import logging
import os
import re
import shutil
import subprocess
import sys
import threading
import time
import unicodedata
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Callable, Dict, Iterator, List, Optional, Sequence, Set, Tuple
from xml.etree import ElementTree

from runtime import (
    DEFAULT_AZURE_MODEL, DEFAULT_OPENAI_MODEL, DEFAULT_REASONING_EFFORT,
    chat_completion_body, configure_process_logging, retry_chat_body,
)

# --- Optional components ---------------------------------------------------
# All of these are optional. Each improves speed or reach; none is required, and
# the tool reports at start-up which ones it found.

try:                                     # faster and stricter workbook reader
    import openpyxl as _openpyxl
except ImportError:
    _openpyxl = None

try:                                     # only used when the model tier is on
    import requests as _requests
except ImportError:
    _requests = None

try:                                     # widens the English word list
    import nltk as _nltk
except ImportError:
    _nltk = None


AGENT_NAME = "Max - wide procurement dataset builder"
AGENT_VERSION = "1.4.0"

# ---------------------------------------------------------------------------
# The agent interface
# ---------------------------------------------------------------------------
# Stages 4 to 6 widen the table with the output of Agents 1, 2 and 3. The column
# names below are that interface, and are asserted against what each agent
# actually wrote: a silently missing column would leave a hole in the Power BI
# model that only shows up in front of the client, so a missing one stops the
# run instead.
#
# Each agent's own file remains the authority on its own analysis. What Max adds
# is the join back onto the transaction row, which the agents cannot do for
# themselves because none of them sees the wide table.

# How a row of an agent's output is traced back to the stage-3 row it came from.
# Agent 1 records the file, sheet and row number it read each line from, so
# feeding it the stage-3 file makes this a faithful pointer into that file. It is
# used rather than Max_Row_Id because Agent 1 emits its own canonical column set
# and does not carry unknown columns through - and the agents are not being
# changed to suit Max.
AGENT_ROW_KEY = "Source_Row_Number"

# Offset between a stage-3 data row and the row number Agent 1 reports for it:
# Agent 1 counts the header as row 1, so its first data row is 2.
AGENT_ROW_OFFSET = 2

AGENT1_REQUIRED: Tuple[str, ...] = (
    "Enriched_Purchase_Description", "Enriched_Description_Short", "Item_Or_Service",
    "AI_Confidence", "Confidence_Band", "Original_Description",
    "Original_Description_Fields", "Detected_Language", "Language_Confidence",
    "Translated_Description", "Translation_Method", "Translation_Coverage",
    "Unresolved_Tokens", "Evidence_Sources", "Evidence_Field_Count", "Match_Tier",
    "Match_Score", "Matched_Source_Systems", "Confidence_Factors", "Source_System",
    "Row_Type", "Is_Duplicate", "Duplicate_Of", "Source_File", "Source_Sheet",
    "Source_Row_Number", "Row_Id", "Run_Id",
    # Which column answered Country. Carried through because Agent 4's
    # Same_Country findings rest on it, and the wide table is where a reader
    # checks whether a country came from a delivery address or a company code.
    "Country", "Country_Source",
)

AGENT2_REQUIRED: Tuple[str, ...] = (
    "AI_Purchase_Group_L5", "AI_Purchase_Group_Id", "AI_Purchase_Group_Confidence",
    "AI_Purchase_Group_Band", "AI_Purchase_Group_Size", "AI_Purchase_Group_Category",
    "AI_Purchase_Group_Cohesion", "AI_Purchase_Group_Naming",
    "AI_Purchase_Group_Is_New", "Agent2_Run_Id",
)

AGENT3_REQUIRED: Tuple[str, ...] = (
    "Standard_item", "Potential_Standard_Match", "Match_Source_Column",
    "Matched_Item_ID", "Matched_Item_Description", "Matched_Item_Supplier",
    "Matched_Item_Source", "Matched_Item_Unit_Price", "Similarity_Score",
    "Match_Band", "Match_Method", "Match_Rationale", "Type_Compatible",
    "Specification_Agreement", "Price_Difference_Percent", "Alternative_Matches",
    "No_Match_Reason", "Closest_Considered_Score", "Closest_Considered_Item_ID",
    "Closest_Considered_Description", "Agent3_Run_Id",
)

# Columns produced by both Max's own stage 3 and by Agent 1. The agent's reading
# wins, because it is the one the client reviewed and the one the downstream
# agents were built on - but only when it is populated, so that a value Max
# already established cannot be blanked by an agent that failed to resolve it.
# Item_Or_Service is the reason this rule exists: Agent 1 decides it with the
# head-word rule and never leaves it blank, while Max's stage 3 guesses.
AGENT_PRECEDENCE_NOTE = ("the agent value wins where it is populated; "
                         "Max's own reading remains in the stage-3 file")

# Supplier identity as Agent 4 resolves it. Added to the wide row so that the
# supplier-consolidation companion file can be joined to it: without this the
# only link would be the raw supplier name, which is the very thing Agent 4
# exists to normalise away.
SUPPLIER_KEY_COLUMN = "Supplier_Key"

# How the client's item catalogue master is named, wherever it has been put.
CATALOGUE_MASTER_GLOB = "*Item*Catalogue*Master*.xls*"

# The resumable stages, in the order they run.
STAGE_ORDER: Tuple[str, ...] = ("wide", "interpret", "agents")

# How each stage is named when a run reports what it kept or where it stopped.
STAGE_LABELS: Dict[str, str] = {
    "wide": "stages 1 and 2",
    "interpret": "stage 3",
    "agents": "the agents",
}

# What a stage leaves half-written if it is interrupted. Deleted when a run
# carries on, so that a truncated table cannot be mistaken for a result. Stages
# recorded as finished are never in here.
PARTIAL_OUTPUTS: Dict[str, Tuple[str, ...]] = {
    "wide": ("max_stage1_sievo_invoice.csv", "max_stage1_sievo_invoice.jsonl",
             "max_stage2_with_po.csv", "max_stage2_with_po.jsonl"),
    "interpret": ("max_stage3_interpreted.csv", "max_stage3_interpreted.jsonl"),
    "agents": ("max_stage4_enriched.csv", "max_stage4_enriched.jsonl",
               "max_stage5_grouped.csv", "max_stage5_grouped.jsonl",
               "max_stage6_standardised.csv", "max_stage6_standardised.jsonl"),
}

# Seconds of quiet from an agent before the build reports that it is still alive.
# Agent 1 translates and embeds in long silent passes, so a build with nothing to
# say for an hour is normal - and indistinguishable from a hung one to whoever is
# watching, who then stops it.
HEARTBEAT_SECONDS = 120

BANNER = f"""
===============================================================================
 Fortum AI-Powered Procurement Analysis
 {AGENT_NAME}
 Prof. Shahab Anbarjafari
===============================================================================
"""

LOGGER = logging.getLogger("max")

# Written into every output row so a result can be traced to the run and to the
# vocabulary version that produced it.
CSV_ENCODING = "utf-8-sig"


def describe_environment() -> Dict[str, bool]:
    return {
        "openpyxl": _openpyxl is not None,
        "requests": _requests is not None,
        "nltk": _nltk is not None,
    }


# ===========================================================================
# Configuration
# ===========================================================================

# List price for the default model, in dollars per million tokens. Overridable
# from the environment because prices are revised and the shared service does
# not have to quote the same rate as the public API.
INPUT_COST_PER_MTOK = 1.25
OUTPUT_COST_PER_MTOK = 10.00

# Default alert threshold offered at the prompt, in dollars.
DEFAULT_SPEND_LIMIT = 25.00


@dataclass
class ModelConfig:
    """Resolved language-model connection details. See .env.example."""

    enabled: bool = False
    backend: str = "openai"
    api_key: str = ""
    base_url: str = "https://api.openai.com/v1"
    model: str = DEFAULT_OPENAI_MODEL
    reasoning_effort: str = DEFAULT_REASONING_EFFORT
    batch_size: int = 20
    timeout: int = 120
    max_requests: int = 0
    spend_limit: float = 0.0         # dollars; 0 means no alert
    input_cost_per_mtok: float = INPUT_COST_PER_MTOK
    output_cost_per_mtok: float = OUTPUT_COST_PER_MTOK

    @property
    def endpoint(self) -> str:
        url = self.base_url.rstrip("/")
        return url if url.endswith("/chat/completions") else f"{url}/chat/completions"


@dataclass
class Settings:
    """Everything the build needs, resolved from arguments and prompts."""

    source_dir: Path
    results_dir: Path
    lexicon_path: Path
    cache_dir: Path

    # Carry every native Maximo and Basware column through to the wide table.
    # Switched off when only the harmonised PO block is wanted.
    native_po_columns: bool = True

    # Free text below this many word characters is not worth interpreting; it is
    # almost always a code or a fragment.
    min_text_length: int = 3

    # A row whose deterministic interpretation resolves less than this share of
    # its content words is a candidate for the model tier.
    interpretation_floor: float = 0.55

    # Run Agents 1 to 4 over the stage-3 file and widen the table with their
    # columns. On by default: the enriched description, the purchase group and the
    # standard-item match are the analysis the client asked for, and a wide table
    # without them is only half the deliverable.
    run_agents: bool = True

    # How long an agent may go without saying anything before it is treated as
    # hung, in seconds. Not a ceiling on how long it may run.
    #
    # This replaces an absolute six-hour limit that did real damage: Agent 1 on the
    # client's full extract was killed at exactly six hours, having worked the
    # whole time and written nothing, which cost that work and every column Agents
    # 2 and 3 would have added afterwards. An absolute limit cannot tell slow
    # progress from a hang, so it eventually kills the run it was meant to protect.
    # Silence can tell them apart: an agent still logging is working however long
    # it takes, and one that has gone quiet is stuck on a prompt or a network call.
    #
    # Two hours, which is deliberately far more than any real pass needs. Agent 3
    # embeds the whole catalogue in one call that prints nothing until it finishes,
    # measured here at 876s of silence for 845,000 items, and that grows with the
    # catalogue and shrinks with the machine. A limit close to the measurement
    # would kill a working agent on a slower laptop, which is the very fault being
    # fixed. Detection speed costs nothing anyway: a genuine hang waits for ever,
    # so catching it late still saves the build, and the heartbeat below is what
    # tells the operator meanwhile that the agent is alive.
    agent_silence_timeout: int = 7200

    # Absolute ceiling, off unless asked for. Available for an unattended run that
    # must finish by a certain time, and best left alone otherwise.
    agent_timeout: Optional[int] = None

    # Rerun every agent even where a result from the same input is already there.
    force_agents: bool = False

    # Carry on from where an interrupted run stopped, rather than repeating the
    # stages it finished. On by default, because repeating them is never what was
    # wanted and the record is only used when it still matches the inputs.
    resume: bool = True

    use_llm: bool = False
    write_jsonl: bool = True
    verbose: bool = False

    # False under --non-interactive, where nothing may block waiting for input.
    interactive: bool = True

    model: ModelConfig = field(default_factory=ModelConfig)


# ---------------------------------------------------------------------------
# Environment handling
# ---------------------------------------------------------------------------

def load_dotenv(path: Path) -> Dict[str, str]:
    """Parse a ``.env`` file into a dictionary.

    Written by hand rather than pulled from a package because the format is
    trivial and this runs on machines where an extra dependency needs approval.
    Later assignments win, matching the shell and python-dotenv.
    """
    values: Dict[str, str] = {}
    if not path.is_file():
        return values

    for raw_line in path.read_text(encoding="utf-8", errors="replace").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        if line.lower().startswith("export "):
            line = line[7:].strip()
        if "=" not in line:
            continue
        key, _, value = line.partition("=")
        key, value = key.strip(), value.strip()
        if len(value) >= 2 and value[0] == value[-1] and value[0] in "\"'":
            value = value[1:-1]
        if key:
            values[key] = value
    return values


def _env_flag(value: Optional[str], default: bool = False) -> bool:
    if value is None:
        return default
    return value.strip().lower() in {"1", "true", "yes", "y", "on", "enable", "enabled"}


def _env_int(value: Optional[str], default: int) -> int:
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return default


def _env_float(value: Optional[str], default: float) -> float:
    """Read a decimal environment variable, tolerating a currency symbol."""
    try:
        return float(str(value).strip().lstrip("$").replace(",", ""))
    except (TypeError, ValueError):
        return default


def resolve_model_config(env: Dict[str, str], use_llm: bool,
                         spend_limit: Optional[float] = None) -> ModelConfig:
    """Select the language-model backend. Mirrors Agents 1 to 4 exactly."""
    config = ModelConfig(enabled=use_llm)
    config.batch_size = max(1, _env_int(env.get("LLM_BATCH_SIZE"), 20))
    config.timeout = max(5, _env_int(env.get("LLM_TIMEOUT"), 120))
    config.max_requests = max(0, _env_int(env.get("LLM_MAX_REQUESTS"), 0))

    if spend_limit is None:
        spend_limit = _env_float(env.get("LLM_SPEND_LIMIT"), DEFAULT_SPEND_LIMIT)
    config.spend_limit = max(0.0, spend_limit)
    config.input_cost_per_mtok = max(
        0.0, _env_float(env.get("LLM_INPUT_COST_PER_MTOK"), INPUT_COST_PER_MTOK))
    config.output_cost_per_mtok = max(
        0.0, _env_float(env.get("LLM_OUTPUT_COST_PER_MTOK"), OUTPUT_COST_PER_MTOK))

    if _env_flag(env.get("AZURE_ENABLE"), False):
        config.backend = "azure"
        config.api_key = (env.get("AZURE_OPENAI_API_KEY") or env.get("AZURE_API_KEY")
                          or env.get("OPENAI_API_KEY") or "")
        config.base_url = (env.get("AZURE_OPENAI_BASE_URL") or env.get("AZURE_BASE_URL")
                           or env.get("BASE_URL")
                           or "https://genai-sharedservice-emea.pwcinternal.com/v1/chat/completions")
        config.model = (env.get("AZURE_OPENAI_MODEL") or env.get("MODEL_NAME")
                        or DEFAULT_AZURE_MODEL)
        config.reasoning_effort = (
            env.get("LLM_REASONING_EFFORT") or DEFAULT_REASONING_EFFORT).strip().lower()
    else:
        # Deliberately does not inherit BASE_URL: that variable points at the
        # shared service on this project, and inheriting it would transmit a
        # personal OpenAI key to an internal endpoint.
        config.backend = "openai"
        config.api_key = env.get("OPENAI_API_KEY") or ""
        config.base_url = env.get("OPENAI_BASE_URL") or "https://api.openai.com/v1"
        config.model = env.get("OPENAI_MODEL") or DEFAULT_OPENAI_MODEL
        config.reasoning_effort = (
            env.get("LLM_REASONING_EFFORT") or DEFAULT_REASONING_EFFORT).strip().lower()

    if config.enabled and not config.api_key:
        LOGGER.warning("Language-model tier requested but no API key was found "
                       "for the %s backend; continuing without it.", config.backend)
        config.enabled = False
    return config


# ===========================================================================
# Text utilities
# ===========================================================================

_CONTROL_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")
_WHITESPACE = re.compile(r"\s+")
_XML_ESCAPES = re.compile(r"_x00[0-9A-Fa-f]{2}_")
_TOKEN = re.compile(r"[^\W_]+", re.UNICODE)

# Mojibake repair table. A UTF-8 string decoded as cp1252 produces these
# sequences; the round trip through latin-1 restores the original bytes. The
# explicit table handles the cases where the round trip is itself lossy.
_MOJIBAKE_MARKERS = ("Ã", "Â", "â€", "Ä\x8d", "Å¡", "Ð")
_MOJIBAKE_LITERAL = {
    "Ã¤": "ä", "Ã¶": "ö", "Ã¥": "å", "Ã„": "Ä", "Ã–": "Ö", "Ã…": "Å",
    "Ã©": "é", "Ã¨": "è", "Ã¼": "ü", "Ã±": "ñ", "Ã¸": "ø", "Ã¦": "æ",
    "â€™": "'", "â€œ": '"', "â€\x9d": '"', "â€“": "-", "â€”": "-", "â€¦": "...",
    "Å‚": "ł", "Å„": "ń", "Å›": "ś", "Å¼": "ż", "Åº": "ź", "Å¾": "ž",
    "Ä…": "ą", "Ä‡": "ć", "Ä™": "ę", "Ã³": "ó", "Â ": " ", "Â": "",
}


def repair_mojibake(text: str) -> str:
    """Undo the common double-encoding damage seen in exported spreadsheets."""
    if not text or not any(marker in text for marker in _MOJIBAKE_MARKERS):
        return text
    try:
        repaired = text.encode("cp1252", errors="strict").decode("utf-8", errors="strict")
        if "\ufffd" not in repaired:
            return repaired
    except (UnicodeEncodeError, UnicodeDecodeError):
        pass
    for damaged, correct in _MOJIBAKE_LITERAL.items():
        text = text.replace(damaged, correct)
    return text


def normalise_text(value: Any) -> str:
    """Reduce any cell value to clean, single-spaced text."""
    if value is None:
        return ""
    if isinstance(value, float):
        # Whole floats are almost always integer keys that a reader widened;
        # rendering 1.0 as "1" keeps join keys comparable across files.
        text = str(int(value)) if value.is_integer() else repr(value)
    elif not isinstance(value, str):
        text = str(value)
    else:
        text = value

    text = _XML_ESCAPES.sub(" ", text)
    text = repair_mojibake(text)
    text = unicodedata.normalize("NFC", text)
    text = _CONTROL_CHARS.sub(" ", text)
    text = _WHITESPACE.sub(" ", text)
    return text.strip()


def fold_accents(text: str) -> str:
    """Strip diacritics for lookup purposes only; never written to output."""
    decomposed = unicodedata.normalize("NFD", text)
    stripped = "".join(ch for ch in decomposed if not unicodedata.combining(ch))
    return (stripped
            .replace("ł", "l").replace("Ł", "L")
            .replace("ø", "o").replace("Ø", "O")
            .replace("æ", "ae").replace("Æ", "Ae")
            .replace("ß", "ss").replace("đ", "d"))


def lookup_key(text: str) -> str:
    """Canonical key for vocabulary and cache lookups."""
    return _WHITESPACE.sub(" ", fold_accents(text).lower()).strip()


def compact_key(value: Any) -> str:
    """Aggressive key for joining identifiers across systems.

    Purchase order numbers are written as ``PO23983000004`` in one system and
    ``po-23983000004`` in another; both reduce to the same key here. Leading
    zeros are preserved because they are significant in ERP numbering.
    """
    return re.sub(r"[^A-Z0-9]", "", normalise_text(value).upper())


def tokenise(text: str) -> List[str]:
    """Split into word tokens, keeping digits attached to their word."""
    return _TOKEN.findall(text)


# Nordic/Polish/German letters that never appear in English output columns.
_FOREIGN_LETTERS = re.compile(r"[äöåÄÖÅąćęłńóśźżĄĆĘŁŃÓŚŹŻõüÜßøæØÆ]")

# Inflectional endings that mark a Finnish, Swedish, Polish or German common
# noun. Kept conservative so English words such as "planning" or "training"
# are not treated as foreign.
_FOREIGN_ENDINGS = (
    "ukset", "uksen", "uksia", "uksella", "ukseen",
    "minen", "mista", "mistä", "ainen", "oinen", "ellinen",
    "työ", "työt", "tyota", "työtä",
    "ningar", "heter", "elser",
    "anie", "enia", "owych", "owie",
    "ungen", "keiten", "schaft",
)

# Frequent procurement words that leak into English columns when the local
# reader treats a Title Case Finnish token as a proper noun. The list is the
# last line of defence, not a vocabulary.
_FOREIGN_TERMS = frozenset("""
kuljetukset kuljetus kuljetuspalvelu huolto huoltotyo huoltotyö kunnossapito
vuokra vuokraus vuokran purku purkutyo purkutyö asbestipurku palvelu palvelut
palvelua korjaus asennus siivous konsultointi koulutus tarkastus hankinta
varaosa varaosat sopimus lasku tilaus työ tyot työt laite laitteet urakka
mittaus kaytto käyttö
arbete underhall underhåll reparation tjanst tjänst tjanster tjänster
hyra uthyrning avtal faktura bestallning beställning
usluga uslugi usługa usługi usuwanie azbestu konserwacja naprawa
wynajem zamowienie zamówienie instalacja
dienstleistung reparatur wartung
""".split())


def is_foreign_common_noun(token: str) -> bool:
    """True when a token is a Finnish/Swedish/Polish/German common noun.

    Proper nouns (places, people, brands) and part numbers return False: they
    are the same in English and must be left alone. Common nouns must not
    appear in the interpreted columns.
    """
    if not token:
        return False
    key = lookup_key(token)
    if not key or any(ch.isdigit() for ch in key):
        return False
    if key in _FOREIGN_TERMS:
        return True
    if _FOREIGN_LETTERS.search(token) and len(key) > 2:
        return True
    return any(key.endswith(ending) and len(key) > len(ending) + 1
               for ending in _FOREIGN_ENDINGS)


def foreign_tokens_in(text: str) -> List[str]:
    """Common nouns in ``text`` that are not English."""
    return [token for token in tokenise(text) if is_foreign_common_noun(token)]


def drop_foreign_common_nouns(text: str) -> str:
    """Remove leftover foreign common nouns so an output column stays English."""
    kept = [token for token in tokenise(text) if not is_foreign_common_noun(token)]
    return sentence_case(" ".join(kept)) if kept else ""


def sentence_case(text: str) -> str:
    """Capitalise the first letter and leave the rest alone.

    ``str.capitalize`` would lower-case the remainder and destroy acronyms such
    as PPE and VAT that legitimately appear inside a description.
    """
    text = text.strip()
    return text[0].upper() + text[1:] if text else ""


def stable_hash(*parts: str) -> str:
    """Short, stable identifier derived from content."""
    return hashlib.sha256("\x1f".join(parts).encode("utf-8")).hexdigest()[:16]


def parse_amount(value: Any) -> Optional[float]:
    """Parse a monetary or quantity cell written in any European convention.

    Returns None rather than zero for unparseable input, because zero is a
    legitimate value and conflating the two would corrupt every sum built on it.
    """
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)

    text = normalise_text(value)
    if not text:
        return None

    negative = text.startswith("(") and text.endswith(")")
    text = re.sub(r"[^\d,.\-]", "", text.strip("()"))
    if not text or text in {"-", ".", ","}:
        return None

    has_comma, has_dot = "," in text, "." in text
    if has_comma and has_dot:
        decimal_sep = "," if text.rfind(",") > text.rfind(".") else "."
        thousands_sep = "." if decimal_sep == "," else ","
        text = text.replace(thousands_sep, "").replace(decimal_sep, ".")
    elif has_comma:
        head, _, tail = text.rpartition(",")
        text = head + tail if (len(tail) == 3 and head) else text.replace(",", ".")

    try:
        amount = float(text)
    except ValueError:
        return None
    return -amount if negative else amount


def format_amount(value: Optional[float]) -> str:
    """Render a number for output without trailing float noise."""
    if value is None:
        return ""
    if abs(value - round(value)) < 1e-9:
        return str(int(round(value)))
    return f"{value:.4f}".rstrip("0").rstrip(".")


# Values that mean "this cell was not filled in". Treating them as data is the
# single most common cause of a join matching the wrong thing.
_NULL_TOKENS = {
    "", "n/a", "na", "n.a.", "none", "null", "nil", "-", "--", "0",
    "unknown", "not defined", "not applicable", "tbd", "#n/a", "nan",
}


def is_blank(value: Any) -> bool:
    """Whether a cell carries no usable value."""
    return normalise_text(value).strip().lower() in _NULL_TOKENS


def line_key(value: Any) -> str:
    """Canonical form of a line number.

    ``1``, ``01`` and ``1.0`` are the same line written by three systems, so all
    three reduce to ``1``. Anything that is not a plain number is compared on
    its compacted text instead.
    """
    text = normalise_text(value)
    if not text or is_blank(text):
        return ""
    try:
        number = float(text.replace(",", "."))
    except ValueError:
        return compact_key(text)
    return str(int(number)) if abs(number - int(number)) < 1e-9 else str(number)


_DOCID_IN_URL = re.compile(r"[?&]docid=([A-Za-z0-9]+)", re.IGNORECASE)
_LEADING_ID = re.compile(r"^([A-Za-z0-9]{16,64})(?:[_.\-]|$)")


def document_id_key(value: Any) -> str:
    """Reduce an opaque document identifier to a comparable form.

    Sievo carries the identifier bare in ``DocumentIdentifier`` and again inside
    the ``docid`` parameter of ``InvoiceLink``; the invoice extract carries it as
    the leading element of ``xml_file_name``. All three spellings are reduced
    here so the three can be compared to each other.
    """
    text = normalise_text(value)
    if not text or is_blank(text):
        return ""

    match = _DOCID_IN_URL.search(text)
    if match:
        return match.group(1).lower()

    # A file name of the form "<id>_15613_15788-1412014052-D.xml".
    stem = text.split("/")[-1].split("\\")[-1]
    match = _LEADING_ID.match(stem)
    if match:
        return match.group(1).lower()

    compact = re.sub(r"[^A-Za-z0-9]", "", text).lower()
    return compact if len(compact) >= 16 else ""


# ===========================================================================
# Table reading
# ===========================================================================
#
# Every reader exposes the same shape: a header list and a factory that yields a
# fresh iterator over the rows. The factory rather than a list keeps memory flat
# on a file of a million rows, and lets the build make more than one pass over a
# source without holding it.

@dataclass
class Table:
    """A single sheet or delimited file, streamed on demand."""

    path: Path
    sheet: str
    headers: List[str]
    open_rows: Callable[[], Iterator[Tuple[int, List[str]]]]

    def iter_rows(self) -> Iterator[Tuple[int, List[str]]]:
        """Yield ``(source_row_number, values)`` with values aligned to headers."""
        width = len(self.headers)
        for row_number, values in self.open_rows():
            if len(values) < width:
                values = values + [""] * (width - len(values))
            elif len(values) > width:
                values = values[:width]
            yield row_number, values

    def iter_records(self) -> Iterator[Tuple[int, Dict[str, str]]]:
        """Yield ``(source_row_number, {header: value})``."""
        for row_number, values in self.iter_rows():
            yield row_number, dict(zip(self.headers, values))

    @property
    def label(self) -> str:
        """Human-readable identity used in logs and diagnostics."""
        if self.sheet in {"", "Sheet1", self.path.stem}:
            return self.path.name
        return f"{self.path.name}:{self.sheet}"


# --- Delimited files -------------------------------------------------------

_ENCODING_CANDIDATES = ("utf-8-sig", "utf-8", "cp1252", "cp1250", "latin-1")


def _detect_encoding(path: Path) -> str:
    """Pick the first encoding that decodes a sample without error.

    The catalogues are Central European and the PO extracts are Nordic, so both
    cp1250 and cp1252 have to be candidates. ``latin-1`` is last because it
    decodes any byte sequence and would mask a better answer.
    """
    sample = path.open("rb").read(1 << 18)
    for encoding in _ENCODING_CANDIDATES:
        try:
            sample.decode(encoding)
        except UnicodeDecodeError:
            continue
        return encoding
    return "latin-1"


def _detect_delimiter(sample: str) -> str:
    """Choose a delimiter by counting candidates outside quoted regions."""
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except csv.Error:
        pass
    counts = {candidate: sample.count(candidate) for candidate in ",;\t|"}
    best = max(counts, key=lambda key: counts[key])
    return best if counts[best] > 0 else ","


def _read_delimited(path: Path) -> List[Table]:
    """Wrap a CSV or TSV file as a single streaming table."""
    encoding = _detect_encoding(path)
    with path.open("r", encoding=encoding, errors="replace", newline="") as handle:
        sample = handle.read(1 << 16)
    delimiter = _detect_delimiter(sample)

    with path.open("r", encoding=encoding, errors="replace", newline="") as handle:
        try:
            header_row = next(csv.reader(handle, delimiter=delimiter))
        except StopIteration:
            return []
    headers = [normalise_text(cell) for cell in header_row]
    if not any(headers):
        return []

    def open_rows() -> Iterator[Tuple[int, List[str]]]:
        with path.open("r", encoding=encoding, errors="replace", newline="") as stream:
            rows = csv.reader(stream, delimiter=delimiter)
            next(rows, None)                       # discard the header
            for index, values in enumerate(rows, start=2):
                yield index, [normalise_text(cell) for cell in values]

    return [Table(path=path, sheet=path.stem, headers=headers, open_rows=open_rows)]


# --- Excel workbooks -------------------------------------------------------

_SPREADSHEET_NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
_PACKAGE_REL_NS = "{http://schemas.openxmlformats.org/package/2006/relationships}"
_DOCUMENT_REL_NS = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"

# Built-in Excel number-format identifiers that denote a date or a time. A cell
# holding a date is stored as a serial number, so without the style lookup an
# order date is emitted as "45103" and every date-based check silently fails.
_BUILTIN_DATE_FORMATS = frozenset({14, 15, 16, 17, 18, 19, 20, 21, 22, 45, 46, 47, 27, 30, 36, 50, 57})


def _column_index(cell_reference: str) -> int:
    """Convert an Excel cell reference such as ``AB12`` to a zero-based column."""
    index = 0
    for char in cell_reference:
        if not char.isalpha():
            break
        index = index * 26 + (ord(char.upper()) - 64)
    return index - 1


def _excel_serial_to_iso(value: Any) -> str:
    """Convert an Excel date serial to an ISO date string.

    Excel's epoch is 1899-12-30 rather than 1900-01-01 because the format
    deliberately reproduces a leap-year bug from Lotus 1-2-3.
    """
    try:
        serial = float(value)
    except (TypeError, ValueError):
        return normalise_text(value)
    if serial <= 0:
        return normalise_text(value)

    from datetime import datetime, timedelta
    moment = datetime(1899, 12, 30) + timedelta(days=serial)
    if abs(serial - int(serial)) < 1e-9:
        return moment.strftime("%Y-%m-%d")
    return moment.strftime("%Y-%m-%d %H:%M:%S")


def _read_xlsx_with_openpyxl(path: Path) -> List[Table]:
    """Stream a workbook through openpyxl in read-only mode."""
    tables: List[Table] = []
    workbook = _openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_names = list(workbook.sheetnames)
    finally:
        workbook.close()

    for sheet_name in sheet_names:
        probe = _openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            header_values: List[str] = []
            for row in probe[sheet_name].iter_rows(min_row=1, max_row=1, values_only=True):
                header_values = [normalise_text(cell) for cell in row]
                break
        finally:
            probe.close()

        if not any(header_values):
            continue

        def open_rows(sheet: str = sheet_name) -> Iterator[Tuple[int, List[str]]]:
            book = _openpyxl.load_workbook(path, read_only=True, data_only=True)
            try:
                for index, row in enumerate(book[sheet].iter_rows(min_row=2, values_only=True),
                                            start=2):
                    yield index, [normalise_text(cell) for cell in row]
            finally:
                book.close()

        tables.append(Table(path=path, sheet=sheet_name,
                            headers=header_values, open_rows=open_rows))
    return tables


def _xlsx_shared_strings(archive: zipfile.ZipFile) -> List[str]:
    """Read the workbook's shared string table."""
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []
    root = ElementTree.fromstring(archive.read("xl/sharedStrings.xml"))
    return ["".join(node.text or "" for node in item.iter(_SPREADSHEET_NS + "t"))
            for item in root.iter(_SPREADSHEET_NS + "si")]


def _xlsx_date_styles(archive: zipfile.ZipFile) -> Set[int]:
    """Identify which cell style indices represent dates."""
    if "xl/styles.xml" not in archive.namelist():
        return set()
    root = ElementTree.fromstring(archive.read("xl/styles.xml"))

    custom_date_formats: Set[int] = set()
    for number_format in root.iter(_SPREADSHEET_NS + "numFmt"):
        code = number_format.attrib.get("formatCode", "").lower()
        # A format is a date format when it references day, month or year
        # placeholders outside of a literal string.
        if re.search(r"(?<!\\)[dmyh]", re.sub(r'"[^"]*"', "", code)):
            custom_date_formats.add(int(number_format.attrib["numFmtId"]))

    date_styles: Set[int] = set()
    cell_formats = root.find(_SPREADSHEET_NS + "cellXfs")
    if cell_formats is None:
        return date_styles
    for style_index, cell_format in enumerate(cell_formats.iter(_SPREADSHEET_NS + "xf")):
        format_id = int(cell_format.attrib.get("numFmtId", 0))
        if format_id in _BUILTIN_DATE_FORMATS or format_id in custom_date_formats:
            date_styles.add(style_index)
    return date_styles


def _iter_xlsx_sheet(payload: bytes, shared: List[str], date_styles: Set[int]) -> Iterator[List[str]]:
    """Yield dense row values from a worksheet part.

    Uses ``iterparse`` and clears each element after use so memory stays flat
    across a very large sheet. Sparse rows are densified against the cell
    references, because a row that omits its empty trailing cells would
    otherwise misalign against the header.
    """
    for _, element in ElementTree.iterparse(io.BytesIO(payload), events=("end",)):
        if element.tag != _SPREADSHEET_NS + "row":
            continue

        cells: Dict[int, str] = {}
        for cell in element.iter(_SPREADSHEET_NS + "c"):
            reference = cell.attrib.get("r", "")
            cell_type = cell.attrib.get("t")
            value_node = cell.find(_SPREADSHEET_NS + "v")
            inline_node = cell.find(_SPREADSHEET_NS + "is")

            if cell_type == "s" and value_node is not None:
                try:
                    text = shared[int(value_node.text)]
                except (ValueError, IndexError, TypeError):
                    text = ""
            elif inline_node is not None:
                text = "".join(node.text or "" for node in inline_node.iter(_SPREADSHEET_NS + "t"))
            elif value_node is not None:
                text = value_node.text or ""
                style = cell.attrib.get("s")
                if cell_type is None and style is not None and int(style) in date_styles:
                    text = _excel_serial_to_iso(text)
            else:
                text = ""

            if reference:
                cells[_column_index(reference)] = normalise_text(text)

        width = max(cells) + 1 if cells else 0
        yield [cells.get(position, "") for position in range(width)]
        element.clear()


def _read_xlsx_with_stdlib(path: Path) -> List[Table]:
    """Read a workbook using only the standard library.

    An .xlsx file is a zip archive of XML parts, so this is entirely supportable
    without a dependency, and having it means a missing package downgrades
    performance rather than stopping the run.
    """
    tables: List[Table] = []
    with zipfile.ZipFile(path) as archive:
        names = archive.namelist()
        if "xl/workbook.xml" not in names:
            return []
        relationships = {
            node.attrib["Id"]: node.attrib["Target"]
            for node in ElementTree.fromstring(
                archive.read("xl/_rels/workbook.xml.rels")).iter(_PACKAGE_REL_NS + "Relationship")
        }
        workbook = ElementTree.fromstring(archive.read("xl/workbook.xml"))
        sheets = [(node.attrib.get("name", "Sheet"),
                   node.attrib.get(_DOCUMENT_REL_NS + "id", ""))
                  for node in workbook.iter(_SPREADSHEET_NS + "sheet")]
        shared = _xlsx_shared_strings(archive)
        date_styles = _xlsx_date_styles(archive)

        for sheet_name, relationship_id in sheets:
            target = relationships.get(relationship_id, "")
            if not target:
                continue
            part = target[1:] if target.startswith("/") else f"xl/{target}"
            if part not in names:
                continue

            headers: List[str] = []
            for candidate in _iter_xlsx_sheet(archive.read(part), shared, date_styles):
                if any(candidate):
                    headers = candidate
                    break
            if not headers:
                continue

            def open_rows(part_name: str = part) -> Iterator[Tuple[int, List[str]]]:
                with zipfile.ZipFile(path) as inner:
                    stream = _iter_xlsx_sheet(inner.read(part_name), shared, date_styles)
                    started = False
                    for index, values in enumerate(stream, start=1):
                        if not started:
                            if any(values):
                                started = True
                            continue
                        yield index, values

            tables.append(Table(path=path, sheet=sheet_name,
                                headers=headers, open_rows=open_rows))
    return tables


def read_table_file(path: Path) -> List[Table]:
    """Read any supported tabular file into one table per sheet."""
    suffix = path.suffix.lower()
    try:
        if suffix in {".csv", ".tsv", ".txt"}:
            return _read_delimited(path)
        if suffix in {".xlsx", ".xlsm"}:
            if _openpyxl is not None:
                return _read_xlsx_with_openpyxl(path)
            return _read_xlsx_with_stdlib(path)
        if suffix == ".xls":
            LOGGER.warning("Legacy .xls is not supported; convert %s to .xlsx", path.name)
            return []
    except Exception as error:
        LOGGER.warning("Could not read %s: %s", path.name, error)
    return []


# ===========================================================================
# Source classification
# ===========================================================================
#
# Which file is which is decided from the header signature rather than the file
# name, so the extracts can be renamed, re-foldered or delivered several at a
# time without touching this code. Each role names the headers that identify it;
# a table is assigned the best-scoring role that clears the threshold.

_ROLE_SIGNATURES: Dict[str, Tuple[str, ...]] = {
    "sievo": ("sourcerowid", "datasource", "spend in eur", "po number",
              "category l1", "erp supplier name", "document number"),
    "invoice": ("invoice_key", "invoice_id", "row_number", "article_name",
                "row_total_excl_vat", "vat_amount", "xml_file_name"),
    "maximo": ("ponum", "polinenum", "line_description", "xpointernalnote",
               "commoditygroup", "orderqty", "unitcost"),
    "basware": ("order number", "po line number", "requisition number",
                "supplier product name", "po net sum company", "main category"),
}

# Below this share of a role's signature the table is left unclassified rather
# than forced into the closest role; a wrong role is worse than none.
_ROLE_THRESHOLD = 0.34


def classify_table(table: Table) -> Tuple[str, float]:
    """Assign a table to a source role, with the score that earned it."""
    present = {lookup_key(header) for header in table.headers if header}
    best_role, best_score = "", 0.0
    for role, signature in _ROLE_SIGNATURES.items():
        hits = sum(1 for name in signature if name in present)
        score = hits / len(signature)
        if score > best_score:
            best_role, best_score = role, score
    if best_score < _ROLE_THRESHOLD:
        return "", best_score
    return best_role, best_score


def discover_tables(source_dir: Path) -> Dict[str, List[Table]]:
    """Read every supported file under the source folder and group by role."""
    grouped: Dict[str, List[Table]] = defaultdict(list)
    paths = sorted(path for path in source_dir.rglob("*")
                   if path.is_file()
                   and path.suffix.lower() in {".csv", ".tsv", ".txt", ".xlsx", ".xlsm", ".xls"}
                   and not path.name.startswith("~$"))

    for path in paths:
        for table in read_table_file(path):
            role, score = classify_table(table)
            if not role:
                LOGGER.info("Skipping %s: does not match a known extract (best score %.2f)",
                            table.label, score)
                continue
            LOGGER.info("%-8s <- %s (%d columns, signature %.0f%%)",
                        role, table.label, len(table.headers), score * 100)
            grouped[role].append(table)
    return grouped


class ColumnMap:
    """Resolves logical field names to the headers a particular table uses.

    Extracts get re-cased, re-spaced and occasionally renamed between deliveries,
    so every field is looked up through a list of accepted spellings rather than
    a single literal. Lookup is on the folded, lower-cased header.
    """

    def __init__(self, headers: Sequence[str], synonyms: Dict[str, Sequence[str]]) -> None:
        index = {lookup_key(header): header for header in headers if header}
        self._resolved: Dict[str, str] = {}
        for field_name, candidates in synonyms.items():
            for candidate in candidates:
                header = index.get(lookup_key(candidate))
                if header:
                    self._resolved[field_name] = header
                    break

    def header(self, field_name: str) -> str:
        return self._resolved.get(field_name, "")

    def has(self, field_name: str) -> bool:
        return field_name in self._resolved

    def get(self, record: Dict[str, str], field_name: str) -> str:
        header = self._resolved.get(field_name)
        return normalise_text(record.get(header, "")) if header else ""

    @property
    def resolved(self) -> Dict[str, str]:
        return dict(self._resolved)


SIEVO_FIELDS: Dict[str, Sequence[str]] = {
    "row_id": ("SourceRowId", "Source Row Id", "RowId"),
    "data_source": ("DataSource", "Data source", "Source system"),
    "document_number": ("Document number", "DocumentNumber", "Document no"),
    "document_line": ("Document line number", "Document line no", "DocumentLineNumber"),
    "document_line_desc": ("Document line desc", "Document line description"),
    "po_number": ("PO number", "PONumber", "Purchase order number", "Order number"),
    "po_line": ("PO line number", "POLineNumber", "Purchase order line number"),
    "po_line_desc": ("PO line desc", "PO line description"),
    "invoice_number": ("Invoice number", "InvoiceNumber", "Invoice no"),
    "document_identifier": ("DocumentIdentifier", "Document identifier"),
    "invoice_link": ("InvoiceLink", "Invoice link"),
    "supplier_name": ("ERP supplier name", "Supplier name", "Supplier"),
    "supplier_number": ("ERP supplier number", "Supplier number", "Supplier code"),
    "spend_eur": ("Spend in EUR", "Spend EUR"),
    "quantity": ("Quantity", "Qty"),
    "material_group_name": ("MaterialGroupName", "Material group name"),
    "category_l4": ("Category L4", "CategoryL4"),
    "project": ("Project",),
}

INVOICE_FIELDS: Dict[str, Sequence[str]] = {
    "invoice_key": ("invoice_key", "invoice key"),
    "invoice_id": ("invoice_id", "invoice id", "invoice number"),
    "row_number": ("row_number", "row number", "line number"),
    "xml_file_name": ("xml_file_name", "xml file name", "file name"),
    "article_id": ("article_id", "article id", "item id"),
    "article_name": ("article_name", "article name", "item name", "description"),
    "free_text": ("free_text", "free text", "text"),
    "quantity_charged": ("quantity_charged", "quantity charged", "quantity"),
    "quantity_delivered": ("quantity_delivered", "quantity delivered"),
    "unit_price_excl_vat": ("unit_price_excl_vat", "unit price excl vat"),
    "unit_price_net": ("unit_price_net", "unit price net"),
    "row_total_excl_vat": ("row_total_excl_vat", "row total excl vat"),
    "row_total_incl_vat": ("row_total_incl_vat", "row total incl vat"),
    "vat_amount": ("vat_amount", "vat amount"),
    "vat_rate": ("vat_rate", "vat rate"),
}

# The two purchase-order systems are mapped onto one harmonised vocabulary so
# that downstream work does not have to branch on which system a row came from.
# Native columns are carried through alongside, unaltered.
MAXIMO_FIELDS: Dict[str, Sequence[str]] = {
    "po_number": ("PONUM",),
    "po_line_number": ("POLINENUM",),
    "header_description": ("DESCRIPTION",),
    "line_description": ("LINE_DESCRIPTION",),
    "item_code": ("ITEMNUM",),
    "category_main": ("COMMODITYGROUP",),
    "category_sub": ("COMMODITY",),
    "quantity": ("ORDERQTY",),
    "unit_cost": ("UNITCOST",),
    "line_cost": ("LINECOST",),
    "total_cost": ("TOTALCOST",),
    "currency": ("CURRENCYCODE",),
    "supplier_name": ("VENDOR",),
    "order_date": ("ORDERDATE",),
    "status": ("STATUS",),
    "order_type": ("POTYPE",),
    "buyer": ("PURCHASEAGENT",),
    "requester": ("XREQUESTEDBY",),
    "payment_terms": ("PAYMENTTERMS",),
    "delivery_terms": ("XDELIVERYTERMS",),
    "contract_reference": ("CONTRACTREFNUM",),
    "offer_reference": ("XOFFERNUM",),
    "frame_contract": ("XFRAMECONTRACT",),
    "internal_note": ("XPOINTERNALNOTE",),
    "terms_conditions": ("XCONTRACTTERMS",),
    "company_code": ("BUYERCOMPANY",),
    "site": ("SITEID",),
}

BASWARE_FIELDS: Dict[str, Sequence[str]] = {
    "po_number": ("Order number",),
    "po_line_number": ("PO line number",),
    "line_description": ("Supplier product name",),
    "item_code": ("Supplier product code", "Item ID"),
    "category_main": ("Main category",),
    "category_sub": ("Sub category",),
    "category_code": ("Category code",),
    "quantity": ("PO line quantity",),
    "line_cost": ("PO net sum company",),
    "currency": ("PO currency company", "PO currency organization"),
    "supplier_name": ("Supplier name",),
    "supplier_code": ("Supplier code",),
    "order_date": ("PO creation date",),
    "status": ("Order status",),
    "line_status": ("Order line status",),
    "order_type": ("Order type",),
    "buyer": ("PO creator",),
    "requester": ("PR owner name", "PR creator"),
    "payment_terms": ("Payment term code",),
    "contract_reference": ("Contract number",),
    "terms_conditions": ("PO terms and conditions",),
    "company_code": ("Company code",),
    "company_name": ("Company name",),
    "cost_center_code": ("Cost center code",),
    "cost_center_name": ("Cost center name",),
    "project_code": ("Project code",),
    "project_name": ("Project name",),
    "account_code": ("Account code",),
    "account_name": ("Account name",),
    "requisition_number": ("Requisition number",),
    "unspsc": ("UNSPSC",),
    "item_type": ("Item type",),
}

# The harmonised block written for whichever system matched.
PO_OUTPUT_FIELDS: Tuple[str, ...] = (
    "po_number", "po_line_number", "header_description", "line_description",
    "item_code", "item_type", "category_main", "category_sub", "category_code",
    "quantity", "unit_cost", "line_cost", "total_cost", "currency",
    "supplier_name", "supplier_code", "order_date", "status", "line_status",
    "order_type", "buyer", "requester", "payment_terms", "delivery_terms",
    "contract_reference", "offer_reference", "frame_contract", "terms_conditions",
    "internal_note", "company_code", "company_name", "cost_center_code",
    "cost_center_name", "project_code", "project_name", "account_code",
    "account_name", "requisition_number", "unspsc", "site",
)


def _title(field_name: str) -> str:
    """Turn a logical field name into an output column name.

    Word boundaries are kept, so ``article_name`` becomes ``Article_Name`` and
    the column reads as ``Invoice_Article_Name``. Running the words together
    would produce ``Invoice_ArticleName``, which is harder to read in a
    spreadsheet header and easy to mistype when referring to the column
    elsewhere in this file.
    """
    return "_".join(part.capitalize() for part in field_name.split("_"))


def _column(prefix: str, field_name: str) -> str:
    """Name an output column, without repeating the prefix inside it.

    The logical field is ``po_number`` and the block prefix is ``PO``, so the
    column is ``PO_Number`` rather than ``PO_Po_Number``.
    """
    lead = f"{prefix.lower()}_"
    if field_name.startswith(lead):
        field_name = field_name[len(lead):]
    return f"{prefix}_{_title(field_name)}"


# ===========================================================================
# Controlled vocabulary
# ===========================================================================

@dataclass
class Lexicon:
    """The shared procurement vocabulary, loaded once and read many times."""

    version: str = "0"
    phrases: Dict[str, Dict[str, str]] = field(default_factory=dict)
    terms: Dict[str, Dict[str, str]] = field(default_factory=dict)
    compound_parts: Dict[str, Dict[str, str]] = field(default_factory=dict)
    service_markers: Set[str] = field(default_factory=set)
    material_markers: Set[str] = field(default_factory=set)
    noise_terms: Set[str] = field(default_factory=set)
    unit_terms: Dict[str, str] = field(default_factory=dict)

    # Phrase lookup is ordered longest-first so that a specific phrase wins over
    # a shorter one contained inside it.
    ordered_phrases: List[Tuple[str, str, str]] = field(default_factory=list)

    @classmethod
    def load(cls, path: Path) -> "Lexicon":
        if not path.is_file():
            LOGGER.warning("No vocabulary at %s; free text will be read with rules only.", path)
            return cls()
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as error:
            LOGGER.warning("Vocabulary at %s is unreadable (%s); continuing without it.",
                           path, error)
            return cls()

        lexicon = cls(
            version=str(payload.get("version", "0")),
            phrases={language: {lookup_key(key): value for key, value in entries.items()}
                     for language, entries in (payload.get("phrases") or {}).items()},
            terms={language: {lookup_key(key): value for key, value in entries.items()}
                   for language, entries in (payload.get("terms") or {}).items()},
            compound_parts={language: {lookup_key(key): value for key, value in entries.items()}
                            for language, entries in (payload.get("compound_parts") or {}).items()},
            service_markers={lookup_key(term) for term in payload.get("service_markers") or []},
            material_markers={lookup_key(term) for term in payload.get("material_markers") or []},
            noise_terms={lookup_key(term) for term in payload.get("noise_terms") or []},
            unit_terms={lookup_key(key): value
                        for key, value in (payload.get("unit_terms") or {}).items()},
        )

        ordered: List[Tuple[str, str, str]] = []
        for language, entries in lexicon.phrases.items():
            for key, value in entries.items():
                ordered.append((key, value, language))
        ordered.sort(key=lambda item: (-len(item[0]), item[0]))
        lexicon.ordered_phrases = ordered

        LOGGER.info("Vocabulary %s loaded: %d phrases, %d terms.",
                    lexicon.version, len(ordered),
                    sum(len(entries) for entries in lexicon.terms.values()))
        return lexicon

    def term(self, key: str) -> Tuple[str, str]:
        """Look a single token up across every language. Returns (english, language)."""
        for language, entries in self.terms.items():
            value = entries.get(key)
            if value:
                return value, language
        return "", ""


# ===========================================================================
# Free-text interpretation
# ===========================================================================
#
# Everything below runs locally. The extraction rules are deliberately
# conservative: each one either finds a well-formed value or reports nothing, so
# that a populated column can be trusted without checking it against the text.

_EMAIL = re.compile(r"\b[\w.+-]+@[\w-]+\.[\w.-]+\b")
_URL = re.compile(r"\bhttps?://\S+|\bwww\.\S+", re.IGNORECASE)
_ISO_DATE = re.compile(r"\b(\d{4})-(\d{1,2})-(\d{1,2})\b")
_EU_DATE = re.compile(r"\b(\d{1,2})\.(\d{1,2})\.(\d{4})\b")
_EU_DATE_SHORT = re.compile(r"\b(\d{1,2})\.(\d{1,2})\.(?!\d)")
_PERCENT = re.compile(r"\b\d+(?:[.,]\d+)?\s?%")

_CURRENCY_WORDS = "eur|sek|pln|nok|dkk|usd|gbp|czk|huf"
_MONEY = re.compile(
    rf"(?:(?P<symbol>[€$£])\s?(?P<after>\d[\d\s\u00a0.,]*)"
    rf"|(?P<before>\d[\d\s\u00a0.,]*)\s?(?P<word>[€$£]|{_CURRENCY_WORDS})\b)",
    re.IGNORECASE)

# Quantity units seen across the Nordic and Polish extracts, mapped to a single
# spelling so that "5 kpl", "5 st" and "5 pcs" become comparable.
_UNIT_CANON = {
    "kpl": "pcs", "st": "pcs", "stk": "pcs", "szt": "pcs", "pcs": "pcs",
    "pc": "pcs", "pce": "pcs", "ea": "pcs", "unit": "pcs", "units": "pcs",
    "kg": "kg", "g": "g", "t": "t", "ton": "t", "tn": "t",
    "m": "m", "mm": "mm", "cm": "cm", "km": "km",
    "m2": "m2", "m3": "m3", "l": "l", "ltr": "l", "litraa": "l",
    "h": "h", "hrs": "h", "hr": "h", "tuntia": "h", "timmar": "h", "godz": "h",
    "pv": "day", "paivaa": "day", "dagar": "day", "dni": "day",
}
_QUANTITY = re.compile(
    r"\b(\d+(?:[.,]\d+)?)\s?(" + "|".join(sorted(_UNIT_CANON, key=len, reverse=True)) + r")\b",
    re.IGNORECASE)

# Two or more numbers joined by x, as in "1250X2500" or "3/4,5 X 1250 X 2500".
_DIMENSION = re.compile(r"\b\d+(?:[.,/]\d+)*(?:\s?[x×]\s?\d+(?:[.,/]\d+)*){1,3}\b", re.IGNORECASE)

_LEAD_TIME_UNITS = {
    "viikko": 7, "viikkoa": 7, "vko": 7, "veckor": 7, "vecka": 7,
    "week": 7, "weeks": 7, "tydzien": 7, "tygodnie": 7,
    "paiva": 1, "paivaa": 1, "pv": 1, "dagar": 1, "dag": 1,
    "day": 1, "days": 1, "dni": 1, "dzien": 1, "arkipaivaa": 1,
    "kuukausi": 30, "kuukautta": 30, "manad": 30, "manader": 30,
    "month": 30, "months": 30, "miesiac": 30,
}
_LEAD_TIME = re.compile(
    r"\b(\d{1,3})\s?(" + "|".join(sorted(_LEAD_TIME_UNITS, key=len, reverse=True)) + r")\b",
    re.IGNORECASE)

# A structured reference: at least two alphanumeric groups joined by hyphens,
# containing a digit. Catches LO1-K570-570-00083 and HAM-LS-661.
_STRUCTURED_REFERENCE = re.compile(r"\b(?=[\w-]*\d)[A-Za-z0-9]{1,10}(?:-[A-Za-z0-9]{1,10}){1,5}\b")

# An English date range such as "06th-12th" satisfies the reference pattern and
# is not a reference.
_ORDINAL_RANGE = re.compile(r"^\d{1,2}(?:st|nd|rd|th)(?:-\d{1,2}(?:st|nd|rd|th))+$",
                            re.IGNORECASE)

# A quotation or order reference introduced by a keyword, in any of the four
# languages the extracts arrive in.
_REFERENCE_KEYWORDS = re.compile(
    r"\b(tarjous|tarjouksen|tarjoukse\w*|offert\w*|offer|quote|quotation|"
    r"tilaus|tilauksen|order|bestallning|zamowienie|"
    r"sopimus|sopimuksen|contract|avtal|umowa)\b[^\w|]{0,4}(?:n[:o]{1,2}\.?\s*)?"
    r"([A-Za-z0-9][A-Za-z0-9/_-]{3,})",
    re.IGNORECASE)
# The gap between the keyword and the number may not contain a pipe: fields are
# joined with " | " before scanning, and without that exclusion the last word of
# one field binds to the first token of the next, reading "Valve order | IT-1"
# as order number IT-1.

# A model or part code: mixed letters and digits, long enough to be meaningful.
_MODEL_CODE = re.compile(r"\b(?=[A-Za-z0-9./-]*\d)(?=[A-Za-z0-9./-]*[A-Za-z])"
                         r"[A-Za-z0-9][A-Za-z0-9./-]{3,}\b")

# Stripped before model-code detection so that a measurement is not mistaken for
# a part number.
_MEASUREMENT_LIKE = re.compile(
    r"\b\d+(?:[.,]\d+)?\s?(?:mm|cm|m|km|kg|g|t|l|ml|m2|m3|v|kv|kw|kwh|w|a|ma|"
    r"bar|pa|kpa|mpa|hz|khz|mhz|gb|tb|mb|%|°c|c)\b|"
    r"\b(?:dn|pn|iso|din|en|sfs)\s?\d+\b", re.IGNORECASE)

_STOPWORDS: Dict[str, Set[str]] = {
    "fi": {"ja", "on", "ei", "se", "etta", "kanssa", "mukaan", "seka", "tai",
           "kun", "myos", "ovat", "ole", "han", "sen", "joka", "vain", "noin"},
    "sv": {"och", "av", "for", "med", "till", "enligt", "samt", "eller", "den",
           "det", "som", "pa", "ar", "vid", "fran", "inte", "har"},
    "pl": {"i", "w", "na", "z", "do", "oraz", "lub", "sie", "jest", "nie",
           "dla", "od", "po", "przy", "za", "to"},
    "en": {"and", "the", "of", "for", "with", "to", "per", "from", "on", "in",
           "at", "by", "or", "as", "is", "are", "be"},
}

_ALL_STOPWORDS: Set[str] = set().union(*_STOPWORDS.values())

# A core English word list, used to tell text that is already English from text
# that is foreign and was not understood. The distinction decides whether a row
# is offered to the model tier, so getting it wrong is expensive in both
# directions: too narrow a list sends perfectly clear English lines to the
# model, and too broad a list lets untranslated Finnish through as though it had
# been read. The list is deliberately weighted towards the vocabulary of
# purchasing, invoicing and industrial maintenance rather than general prose,
# and it is extended at load time by every English target in the controlled
# vocabulary and, when the corpus is installed, by NLTK's word list.
_ENGLISH_CORE: Set[str] = set("""
a able about above accept access according account accounting accrual acquire
across activity actual adapter adaptor add additional address adjust administration
advance advisory after agreed agreement air alarm all allowance along also alternative
aluminium amount analysis annual another answer any application approval approved
april area arrange article assembly assessment asset assistance associated assurance
august authority automatic available average award back badge bag balance bank bar
base basic basis batch bathroom battery bearing before below belt bench between bid
bill black block blue board body bolt bonus book booking bottle bottom box bracket
brake branch brand breakdown bridge bring broadband budget build building bulk bundle
business but buy buyer cabin cabinet cable calculation calibration call camera cancel
cap capacity capital car carbon card care cargo carriage carrier case cash catalogue
category cell cement central centre certificate certification chain chair change
channel charge charger check chemical circuit civil claim clamp cleaning clear client
clip clothing coating code coil cold collection colour column combined commission
commissioning committee commodity communication company complete compliance component
composite compressor computer concrete condition conference configuration connection
consultancy consultant consulting consumable container content continuous contract
contractor control conversion conveyor cooling copper copy core corner corporate
correction cost council counter coupling cover crane credit crew cross current custom
customer cut cycle cylinder daily damage data date day dealer december decision deck
declaration decommissioning deduction deep defect definition delivery demand demolition
department deposit depot design desk detail detection development device diagnostic
diameter diesel digital dimension direct director disc discharge discount dismantling
dispatch display disposal distance distribution district diverse document domestic
door double down download drain drawing drilling drive driver drum dry due duct duty
early earth economy edge education effect efficiency electric electrical electricity
electronic element email emergency emission employee enclosure end energy engine
engineer engineering enquiry ensure entry environment environmental equipment
estimate european evaluation event examination exchange excluding execution exhaust
existing exit expense expenses expert export exposure express extended external extra
fabric fabrication facility factor factory failure fan february fee female fence field
file filter final finance financial finish fire firm first fitting fixed fixing flange
flat fleet flexible flight flights floor flow fluid foam follow food foot force
forecast form foundation frame framework free freight frequency fresh from fuel full
function furniture fuse gas gasket gate gauge gear general generation generator glass
glove goods grade graphic grease green grid grinding ground group guard guarantee
guide handle handling hardware hazardous head header health heat heater heating heavy
height help high hire hire holder hole holiday home hook hose hospital hotel hour
hourly house housing hydraulic ice identification image impact implementation import
improvement inch include including income increase independent index indirect
individual induction industrial industry information infrastructure initial injection
inspection install installation instrument insulation insurance integration interface
interim internal international internet interview inventory investigation invoice iron
issue item january job joint journal july june keep key keyboard kit knife label
labour laboratory ladder lamp land landfill language laptop large laser last late
layout lead leak lease leasing legal length lens level licence license lift light
lighting lime limit line liner link liquid list load loading loan local location lock
logistics long loss low lubricant machine machinery magnetic main maintenance major
male management manager manual manufacture manufacturer march marine mark market mass
master match material may measure measurement mechanical media medical medium meeting
member membrane metal meter method metre microphone middle mileage mill mineral
minimum mining minor mixer mobile model modification module monitor monitoring month
monthly mortar motor mount mounting mouse move multiple national natural network new
next night noise nominal north note notice november nozzle number nut object
observation october off offer office officer offshore oil online open operating
operation operator option order organisation original other outlet output outsourcing
overhaul overhead overtime package packaging pad paint painting pallet panel paper
parking part partial particle partner parts passenger patch payment peak penalty
pension per performance period permit person personal personnel petrol phase phone
photo physical pick picture piece pile pilot pin pipe pipeline piping piston place
plan planning plant plastic plate platform plug plumbing point pole police policy
polish pollution pool port portable position post power practice preparation pressure
prevention price primary print printer printing private procedure process procurement
product production professional profile programme project promotion property
protection provision public pull pump purchase purchasing pure push quality quantity
quarter quarterly quote quotation rack radio rail railway rain range rate rating raw
reactor real rebate receipt receiver recharge reclamation recommendation reconditioning
record recovery recruitment recycling reduction reference refrigeration refurbishment
refuse regional register regulation regulator reimbursement reinforcement relay release
remote removal renewal rent rental repair replacement report request requirement
rescue research reserve reset residual resin resource response rest restoration
result retail retrofit return review revision rig rig right ring risk river road
robot rock rod roll roof room rope rotor round rounding route routine rubber rubbish
running safety sale sample sand sanitary scaffold scaffolding scale scan scanner
schedule scheme school science scope scrap screen screw sea seal sealing search season
seat second secondary section sector security segment selection self seminar sensor
separate september series service servicing session set setting shaft share sheet
shelf shield shipping shop short shutdown side sign signal silicone simple single site
size skid skill sleeve slide small smart social socket software solar solid solution
solvent sound source south space spare special specialist specification speed spill
spindle spare split spool spray spreader spring stack staff stage stainless stair
stand standard start station steam steel step sterile stock stop storage store
straight strap strategy street stress strip structural structure study sub subcontract
subscription substation substitute supervision supplier supply support surface survey
suspension sweep switch switchgear system table tablet tank tape target task tax team
technical technician technology telecom telephone temporary terminal termination test
testing text thermal thermometer thickness third thread three through ticket tie tight
tile time tin tip tool toolkit top torque total tower town track traffic trailer
training transfer transformer transition transmission transport transportation travel
tray treatment trial trim trip truck tube tuning tunnel turbine turn twin two type
tyre unit universal upgrade urgent usage use used user utility vacuum valve van
variable vat vehicle vendor vent ventilation verification version vertical vessel
video view vinyl visit visual voltage volume voucher wage wall warehouse warranty
waste watch water weather web week weekly weight weld welding well west wheel white
wide width winch wind window wire wireless wood work worker working workshop workwear
wrap yard year yearly yellow zinc zone
""".split())

# A word shared by two languages says nothing about which one is in front of
# you. Swedish "for" and English "for", Polish "to" and English "to" were enough
# on their own to label an English line as Scandinavian, so only the words that
# belong to exactly one language are allowed to vote.
_DISCRIMINATING_STOPWORDS: Dict[str, Set[str]] = {
    language: {word for word in words
               if sum(word in other for other in _STOPWORDS.values()) == 1}
    for language, words in _STOPWORDS.items()
}

# Signals that place a purchase in a recognisable class. Ordered: the first
# group whose markers appear decides, so the more specific groups come first.
_INTENT_RULES: Tuple[Tuple[str, Tuple[str, ...]], ...] = (
    ("Freight and delivery",
     ("freight", "rahti", "toimituskulu", "frakt", "delivery cost", "shipping",
      "kuljetus", "transport", "spedycja", "postage", "courier")),
    ("Travel and accommodation",
     ("travel", "matka", "flight", "flights", "hotel", "accommodation", "b&b",
      "car hire", "taxi", "resa", "podroz", "majoitus", "lento")),
    ("Rental and leasing",
     ("rental", "rent", "vuokra", "vuokraus", "hyra", "uthyrning", "lease",
      "leasing", "wynajem", "hire")),
    ("Training and certification",
     ("training", "koulutus", "utbildning", "kurs", "kurssi", "szkolenie",
      "certification", "course", "seminar", "workshop")),
    ("Consulting and advisory",
     ("consulting", "consultancy", "consultant", "konsultointi", "konsultation",
      "konsult", "doradztwo", "advisory", "audit", "study", "utredning",
      "inventering", "survey", "promemoria")),
    ("Maintenance and repair",
     ("maintenance", "huolto", "kunnossapito", "underhall", "konserwacja",
      "repair", "korjaus", "reparation", "naprawa", "service", "overhaul",
      "kalibrointi", "calibration", "besiktning", "inspection", "tarkastus")),
    ("Installation and construction",
     ("installation", "asennus", "montaz", "construction", "rakennus",
      "byggnad", "erection", "pystytys", "commissioning", "perustus")),
    ("Licences and subscriptions",
     ("licence", "license", "lisenssi", "licens", "licencja", "subscription",
      "tilaus", "software", "ohjelmisto", "programvara", "oprogramowanie",
      "saas", "support agreement")),
    ("Utilities and energy",
     ("electricity", "sahko", "el", "gas", "water", "vesi", "heat", "lampo",
      "district heating", "kaukolampo", "fjarrvarme", "energy")),
    ("Spare parts and materials",
     ("spare", "varaosa", "reservdel", "czesc", "part", "parts", "component",
      "material", "supplies", "tarvike", "consumable", "seal", "tiiviste",
      "valve", "venttiili", "bearing", "laakeri", "cable", "kaapeli")),
    ("Fees and charges",
     ("fee", "charge", "maksu", "avgift", "oplata", "booking fee", "rounding",
      "surcharge", "commission", "hallinnointi")),
)

_ATTACHMENT_MARKERS = ("liite", "liitteessa", "liitteena", "attached", "attachment",
                       "bilaga", "bifogad", "zalacznik", "enclosed", "ohessa")
_STOCK_MARKERS = ("varasto", "varastoon", "varastossa", "lager", "stock",
                  "magazyn", "warehouse", "inventory")
_URGENCY_MARKERS = ("kiireellinen", "urgent", "asap", "bradskande", "pilne",
                    "pikaisesti", "immediately", "heti")
_TAX_MARKERS = ("alv", "vat", "moms", "vat", "podatek", "mwst", "tax")


def _nltk_english_words() -> Set[str]:
    """NLTK's English word list, when the corpus has been downloaded.

    Optional. It widens the core list to ordinary English prose, which matters
    for the UK invoice descriptions. Absence costs a little recall on English
    text and nothing else, so a missing corpus is not worth a warning.
    """
    if _nltk is None:
        return set()
    try:
        from nltk.corpus import words as nltk_words
        return {word.lower() for word in nltk_words.words() if word.isalpha()}
    except Exception:
        return set()


@dataclass
class Interpretation:
    """Everything read from one row's free text."""

    text: str = ""
    sources: str = ""
    language: str = ""
    language_confidence: float = 0.0
    description: str = ""
    method: str = "none"
    confidence: float = 0.0
    item_or_service: str = ""
    intent: str = ""
    keywords: str = ""
    resolved_share: float = 0.0
    token_count: int = 0

    quantity: str = ""
    unit: str = ""
    amount: str = ""
    currency: str = ""
    dimensions: str = ""
    model_codes: str = ""
    references: str = ""
    dates: str = ""
    emails: str = ""
    domains: str = ""
    lead_time_days: str = ""
    percentages: str = ""

    mentions_tax: str = ""
    mentions_attachment: str = ""
    mentions_stock: str = ""
    mentions_urgency: str = ""

    def as_columns(self) -> Dict[str, str]:
        return {
            "Text_Sources": self.sources,
            "Text_Interpreted": self.text,
            "Text_Language": self.language,
            "Text_Language_Confidence": f"{self.language_confidence:.2f}" if self.language else "",
            "Text_Token_Count": str(self.token_count) if self.token_count else "",
            "Interpreted_Description": self.description,
            "Interpretation_Method": self.method,
            "Interpretation_Confidence": f"{self.confidence:.2f}" if self.text else "",
            "Interpretation_Resolved_Share": f"{self.resolved_share:.2f}" if self.text else "",
            "Item_Or_Service": self.item_or_service,
            "Purchase_Intent": self.intent,
            "Keywords": self.keywords,
            "Extracted_Quantity": self.quantity,
            "Extracted_Unit": self.unit,
            "Extracted_Amount": self.amount,
            "Extracted_Currency": self.currency,
            "Extracted_Dimensions": self.dimensions,
            "Extracted_Model_Codes": self.model_codes,
            "Extracted_References": self.references,
            "Extracted_Dates": self.dates,
            "Extracted_Emails": self.emails,
            "Extracted_Domains": self.domains,
            "Extracted_Lead_Time_Days": self.lead_time_days,
            "Extracted_Percentages": self.percentages,
            "Mentions_Tax": self.mentions_tax,
            "Mentions_Attachment": self.mentions_attachment,
            "Mentions_Stock": self.mentions_stock,
            "Mentions_Urgency": self.mentions_urgency,
        }

    @staticmethod
    def columns() -> List[str]:
        return list(Interpretation().as_columns().keys())


class TextInterpreter:
    """Reads a free-text field into structured columns without a model.

    The work splits in two. The extractors pull out values that have an
    unambiguous written form — amounts, quantities, dates, addresses, references
    — and are exact by construction. The reader then renders the remaining prose
    into English through the controlled vocabulary, reporting the share of
    content words it managed to resolve so that a weak reading is visible rather
    than merely wrong.
    """

    def __init__(self, lexicon: Lexicon, settings: Settings) -> None:
        self.lexicon = lexicon
        self.settings = settings
        self._cache: Dict[str, Interpretation] = {}
        self.english_words = self._english_vocabulary(lexicon)
        self.discriminating_terms = self._discriminating_terms(lexicon)

    @staticmethod
    def _english_vocabulary(lexicon: Lexicon) -> Set[str]:
        """The English side of the vocabulary, used to recognise English input.

        Every target in the vocabulary is by definition an English procurement
        word, so the file that translates into English also describes what
        English looks like. Without this, an English line resolves nothing and
        is scored as badly understood when in fact it needed no work at all.
        """
        words: Set[str] = set()
        for mapping in (lexicon.phrases, lexicon.terms, lexicon.compound_parts):
            for entries in mapping.values():
                for value in entries.values():
                    words.update(tokenise(lookup_key(value)))
        for marker in lexicon.service_markers | lexicon.material_markers:
            words.update(tokenise(marker))
        for value in lexicon.unit_terms.values():
            words.update(tokenise(lookup_key(value)))
        words |= _STOPWORDS["en"]
        words |= _ENGLISH_CORE
        words |= _nltk_english_words()
        return {word for word in words if word}

    @staticmethod
    def _discriminating_terms(lexicon: Lexicon) -> Dict[str, Set[str]]:
        """Vocabulary entries that belong to exactly one language.

        'transport', 'material' and 'system' are spelled the same in three of
        the four languages here and so identify none of them.
        """
        counts: Counter = Counter()
        for entries in lexicon.terms.values():
            counts.update(entries.keys())
        return {language: {term for term in entries if counts[term] == 1}
                for language, entries in lexicon.terms.items()}

    # -- entry point --------------------------------------------------------

    def interpret(self, describing: str, supporting: str, sources: str) -> Interpretation:
        """Interpret one row's free text, memoised on that text.

        Distinct strings are far fewer than rows in every extract seen so far,
        so caching here is what keeps a million-row run affordable.
        """
        describing = normalise_text(describing)
        supporting = normalise_text(supporting)
        combined = " | ".join(part for part in (describing, supporting) if part)
        if len(re.sub(r"\W", "", combined)) < self.settings.min_text_length:
            return Interpretation(text=combined, sources=sources)

        cached = self._cache.get((describing, supporting))
        if cached is not None:
            # Sources vary per row while the reading does not; copy and relabel.
            return Interpretation(**{**cached.__dict__, "sources": sources})

        result = self._interpret_uncached(describing, supporting)
        result.sources = sources
        self._cache[(describing, supporting)] = result
        return result

    def _interpret_uncached(self, describing: str, supporting: str) -> Interpretation:
        combined = " | ".join(part for part in (describing, supporting) if part)
        result = Interpretation(text=combined)

        # Values are drawn from everything the row carries: a lead time or a
        # quotation number is just as real in the buyer's note as in the line
        # description, and often only appears there.
        self._extract_values(combined, lookup_key(combined), result)

        # The description is built from the descriptive fields alone, and from
        # the prose within them. Codes, references and addresses have already
        # been captured in their own columns, and leaving them in produces a
        # description like "leasing contract ham".
        subject = describing or supporting
        language, confidence = self._detect_language(lookup_key(subject))
        result.language, result.language_confidence = language, confidence

        prose = self._prose(subject)
        english, resolved, total = self._render_english(
            lookup_key(prose), self._proper_nouns(prose) if language == "en" else set())
        result.resolved_share = resolved / total if total else 0.0
        result.token_count = total
        # Foreign common nouns stay on the reading so the model can translate
        # them; they are stripped from the published columns if the model is
        # off or does not answer.
        result.description = sentence_case(english)
        result.method = "vocabulary" if resolved else ("passthrough" if english else "none")

        # Type and intent are judged on everything the row carries, since a note
        # saying "to stock" or "urgent" bears on both. Keywords come from the
        # rendered description, so that the indexed terms are the English ones.
        folded_all = lookup_key(combined)
        result.item_or_service = self._classify_type(folded_all, english)
        result.intent = self._classify_intent(folded_all, english)
        result.keywords = self._keywords(english)
        result.confidence = self._confidence(result)
        return result

    # -- extraction ---------------------------------------------------------

    def _extract_values(self, text: str, folded: str, result: Interpretation) -> None:
        """Pull the values that have an unambiguous written form."""
        emails = sorted({match.group(0).lower() for match in _EMAIL.finditer(text)})
        result.emails = "; ".join(emails)
        result.domains = "; ".join(sorted({address.split("@")[-1] for address in emails}))

        # Addresses and links are removed before the remaining scans so that a
        # domain or a query string cannot be read as a code or a date.
        scrubbed = _URL.sub(" ", _EMAIL.sub(" ", text))

        quantities = _QUANTITY.findall(scrubbed)
        if quantities:
            amount, unit = quantities[0]
            result.quantity = format_amount(parse_amount(amount))
            result.unit = _UNIT_CANON.get(unit.lower(), unit.lower())

        money = self._extract_money(scrubbed)
        if money:
            result.amount, result.currency = money

        dimensions = sorted({match.group(0).replace(" ", "")
                             for match in _DIMENSION.finditer(scrubbed)})
        result.dimensions = "; ".join(dimensions[:5])

        result.dates = "; ".join(self._extract_dates(scrubbed)[:6])
        result.references = "; ".join(self._extract_references(scrubbed)[:6])
        result.model_codes = "; ".join(self._extract_model_codes(scrubbed)[:6])
        result.percentages = "; ".join(
            sorted({match.group(0).replace(" ", "") for match in _PERCENT.finditer(scrubbed)})[:4])

        lead_days = self._extract_lead_time(folded)
        result.lead_time_days = str(lead_days) if lead_days else ""

        result.mentions_tax = "Yes" if self._mentions(folded, _TAX_MARKERS) else ""
        result.mentions_attachment = "Yes" if self._mentions(folded, _ATTACHMENT_MARKERS) else ""
        result.mentions_stock = "Yes" if self._mentions(folded, _STOCK_MARKERS) else ""
        result.mentions_urgency = "Yes" if self._mentions(folded, _URGENCY_MARKERS) else ""

    @staticmethod
    def _mentions(folded: str, markers: Sequence[str]) -> bool:
        tokens = set(tokenise(folded))
        return any(marker in tokens or (" " in marker and marker in folded)
                   for marker in markers)

    @staticmethod
    def _extract_money(text: str) -> Optional[Tuple[str, str]]:
        """Return the first well-formed money value as (amount, currency)."""
        symbols = {"€": "EUR", "$": "USD", "£": "GBP"}
        for match in _MONEY.finditer(text):
            raw = match.group("after") or match.group("before") or ""
            token = match.group("symbol") or match.group("word") or ""
            amount = parse_amount(raw)
            if amount is None:
                continue
            currency = symbols.get(token, token.upper())
            return format_amount(amount), currency
        return None

    @staticmethod
    def _extract_dates(text: str) -> List[str]:
        """Collect dates, normalised to ISO where the year is known.

        Nordic internal notes routinely write a bare "11.3." meaning a day in the
        current working year. The day and month are kept as written in that case
        rather than guessed into a year that may be wrong.
        """
        found: List[str] = []
        for match in _ISO_DATE.finditer(text):
            year, month, day = match.groups()
            if 1 <= int(month) <= 12 and 1 <= int(day) <= 31:
                found.append(f"{int(year):04d}-{int(month):02d}-{int(day):02d}")
        for match in _EU_DATE.finditer(text):
            day, month, year = match.groups()
            if 1 <= int(month) <= 12 and 1 <= int(day) <= 31:
                found.append(f"{int(year):04d}-{int(month):02d}-{int(day):02d}")
        for match in _EU_DATE_SHORT.finditer(text):
            day, month = match.groups()
            if 1 <= int(month) <= 12 and 1 <= int(day) <= 31:
                found.append(f"--{int(month):02d}-{int(day):02d}")
        # Sorted for determinism, de-duplicated because a note often repeats one.
        return sorted(set(found))

    @staticmethod
    def _extract_references(text: str) -> List[str]:
        """Collect quotation, order and contract references.

        A reference introduced by a keyword is reported with that keyword, since
        knowing a number is a contract rather than a quotation is most of its
        value. The bare forms of those same numbers are then suppressed so the
        column does not carry each reference twice.
        """
        found: List[str] = []
        named: Set[str] = set()

        for match in _REFERENCE_KEYWORDS.finditer(text):
            candidate = match.group(2).strip(".,;:")
            if any(ch.isdigit() for ch in candidate) and len(candidate) >= 4:
                found.append(f"{match.group(1).lower()}:{candidate}")
                named.add(candidate)

        for match in _STRUCTURED_REFERENCE.finditer(text):
            candidate = match.group(0)
            if candidate in named or len(candidate) < 7:
                continue
            # A date written 12-05-2026, and an English date range written
            # 06th-12th, both satisfy the pattern and neither is a reference.
            if re.fullmatch(r"\d{1,4}(?:-\d{1,4}){1,2}", candidate):
                continue
            if _ORDINAL_RANGE.match(candidate):
                continue
            # At least one run of two or more letters, which excludes a number
            # broken up by hyphens.
            if not re.search(r"[A-Za-z]{2,}", candidate):
                continue
            found.append(candidate)
        return sorted(set(found))

    @staticmethod
    def _extract_model_codes(text: str) -> List[str]:
        """Collect part and model codes, with measurements set aside first.

        Measurements and dimensions are removed before the scan because both
        satisfy the pattern of a part number and neither is one; they are
        already reported in their own columns.
        """
        stripped = _DIMENSION.sub(" ", _MEASUREMENT_LIKE.sub(" ", text))
        found = set()
        for match in _MODEL_CODE.finditer(stripped):
            candidate = match.group(0).strip("./-")
            if len(candidate) < 4 or _ORDINAL_RANGE.match(candidate):
                continue
            digits = sum(ch.isdigit() for ch in candidate)
            letters = sum(ch.isalpha() for ch in candidate)
            # Both a letter and a digit are required, which excludes bare
            # quantities and bare words alike.
            if digits and letters:
                found.add(candidate)
        return sorted(found)

    @staticmethod
    def _extract_lead_time(folded: str) -> int:
        """Longest stated lead time in days, or zero when none is stated."""
        best = 0
        for match in _LEAD_TIME.finditer(folded):
            count = int(match.group(1))
            multiplier = _LEAD_TIME_UNITS.get(match.group(2).lower(), 0)
            best = max(best, count * multiplier)
        return best

    # -- language and rendering --------------------------------------------

    def _detect_language(self, folded: str) -> Tuple[str, float]:
        """Identify the language from vocabulary and stop-word hits.

        Scoring rather than a hard rule, because procurement free text mixes
        languages inside a single field far more often than prose does.
        """
        tokens = tokenise(folded)
        if not tokens:
            return "", 0.0

        scores: Dict[str, float] = defaultdict(float)
        for language, stopwords in _DISCRIMINATING_STOPWORDS.items():
            scores[language] += sum(1.5 for token in tokens if token in stopwords)
        for language, entries in self.discriminating_terms.items():
            scores[language] += sum(1.0 for token in tokens if token in entries)
        for key, _, language in self.lexicon.ordered_phrases:
            if key in folded:
                scores[language] += 2.0

        # English words carry less weight than a foreign term, because several
        # of them are also spelled that way in the other three languages.
        scores["en"] += sum(0.75 for token in tokens
                            if token in self.english_words and token not in _ALL_STOPWORDS)

        # Nothing voted. The honest answer is that the language is unknown, not
        # that it is English: the text handed in here has already been folded,
        # so a Swedish word is as plainly ASCII as an English one and an
        # alphabet test would call "Tillstandsansokan" English. Saying nothing
        # keeps the reading's confidence low, which is what routes the line to
        # the model tier where it belongs.
        if not scores or max(scores.values()) == 0:
            return "", 0.0

        best = max(scores, key=lambda language: scores[language])
        total = sum(scores.values()) or 1.0
        return best, min(1.0, scores[best] / total)

    @staticmethod
    def _prose(text: str) -> str:
        """Remove the parts of the text that are identifiers rather than words."""
        cleaned = _URL.sub(" ", _EMAIL.sub(" ", text))
        cleaned = _DIMENSION.sub(" ", cleaned)
        cleaned = _STRUCTURED_REFERENCE.sub(" ", cleaned)
        cleaned = _MODEL_CODE.sub(" ", cleaned)
        return cleaned

    @staticmethod
    def _proper_nouns(prose: str) -> Set[str]:
        """Capitalised words other than the first, taken to be names.

        Only consulted for text already identified as English, where an unknown
        capitalised word is a place, a brand or a person rather than a word that
        failed to translate. Without this, "Car Hire Charges for Teesside" is
        scored as one word short of understood and sent to the model to be told
        that Teesside is a place.
        """
        names: Set[str] = set()
        for position, match in enumerate(_TOKEN.finditer(prose)):
            word = match.group(0)
            # A Title Case Finnish common noun such as Kuljetukset is not a
            # name. Treating it as one is how untranslated words used to survive
            # into Interpreted_Description on mixed-language rows.
            if position and word[:1].isupper() and not is_foreign_common_noun(word):
                names.add(lookup_key(word))
        return names

    def _render_english(self, folded: str, names: Set[str]) -> Tuple[str, int, int]:
        """Render the text in English. Returns (english, resolved, total)."""
        working = folded
        rendered: List[str] = []

        # Phrases first, longest first, so a multi-word entry is not broken up
        # by the single-word pass that follows.
        resolved_phrase_tokens = 0
        for key, value, _ in self.lexicon.ordered_phrases:
            if key and key in working:
                working = working.replace(key, " \u0000 ")
                rendered.append(value)
                resolved_phrase_tokens += len(key.split())

        resolved, total = resolved_phrase_tokens, resolved_phrase_tokens
        for token in tokenise(working):
            if token == "\u0000":
                continue
            if token in self.lexicon.noise_terms:
                continue
            if token.isdigit():
                continue
            total += 1

            english, _ = self.lexicon.term(token)
            if english:
                rendered.append(english)
                resolved += 1
                continue

            unit = self.lexicon.unit_terms.get(token)
            if unit:
                rendered.append(unit)
                resolved += 1
                continue

            decomposed = self._decompose(token)
            if decomposed:
                rendered.append(decomposed)
                resolved += 1
                continue

            # A word that is already English needs no translation, so it counts
            # as resolved. Judging otherwise would score a perfectly clear
            # English line as badly understood and send it to the model tier for
            # no reason, which is most of what the model tier costs. Membership
            # of the word list is required: inferring it from the detected
            # language would let a whole untranslated Finnish line through as
            # understood on the strength of one ambiguous token.
            if is_foreign_common_noun(token):
                # Kept so the model can see what was there, but not counted as
                # understood, so mixed-language lines fall through to the model
                # instead of being accepted as English.
                if len(token) > 2:
                    rendered.append(token)
                continue
            if token in self.english_words or token in names:
                if len(token) > 2 or token in _STOPWORDS["en"]:
                    rendered.append(token)
                resolved += 1
                continue

            # Unknown. Kept rather than dropped so nothing is silently lost,
            # but not counted as understood.
            if len(token) > 2 and not any(ch.isdigit() for ch in token):
                rendered.append(token)

        english = " ".join(dict.fromkeys(word for word in " ".join(rendered).split() if word))
        return english, resolved, total

    def _decompose(self, token: str) -> str:
        """Split a Nordic compound into known parts.

        'asbestipurkutyo' is one token to a tokeniser and three ideas to a
        reader; splitting it is what lets the vocabulary cover a language that
        forms new words by joining old ones.
        """
        for language in ("fi", "sv"):
            parts = self.lexicon.compound_parts.get(language) or {}
            if not parts:
                continue
            pieces = self._split_compound(token, parts)
            if pieces:
                return " ".join(piece for piece in pieces if piece)
        return ""

    @staticmethod
    def _split_compound(token: str, parts: Dict[str, str], depth: int = 0) -> List[str]:
        """Greedy longest-prefix decomposition, at most three pieces deep."""
        if not token:
            return []
        if depth >= 3:
            return []
        for length in range(len(token), 2, -1):
            head = token[:length]
            if head not in parts:
                continue
            tail = token[length:]
            if not tail:
                return [parts[head]]
            rest = TextInterpreter._split_compound(tail, parts, depth + 1)
            if rest:
                return [parts[head]] + rest
        return []

    # -- classification -----------------------------------------------------

    def _classify_type(self, folded: str, english: str) -> str:
        """Decide whether the line buys a thing or an activity."""
        tokens = set(tokenise(english)) | set(tokenise(folded))
        service = len(tokens & self.lexicon.service_markers)
        material = len(tokens & self.lexicon.material_markers)
        if service > material:
            return "Service"
        if material > service:
            return "Material"
        return "Unclassified" if not tokens else ""

    @staticmethod
    def _classify_intent(folded: str, english: str) -> str:
        """Place the purchase in a recognisable class, or leave it unnamed."""
        haystack = f"{folded} {lookup_key(english)}"
        tokens = set(tokenise(haystack))
        for label, markers in _INTENT_RULES:
            for marker in markers:
                if " " in marker:
                    if marker in haystack:
                        return label
                elif marker in tokens:
                    return label
        return ""

    def _keywords(self, english: str) -> str:
        """The content words worth indexing, in a stable order."""
        words = [word for word in tokenise(lookup_key(english))
                 if len(word) > 2 and word not in _ALL_STOPWORDS
                 and word not in self.lexicon.noise_terms
                 and not word.isdigit()]
        return "; ".join(sorted(dict.fromkeys(words))[:12])

    @staticmethod
    def _confidence(result: Interpretation) -> float:
        """How much of the row's text was actually understood.

        Built from the share of content words resolved, lifted by each
        structured value that was extracted cleanly, because a row where the
        quantity, the amount and the reference all came out is well understood
        even when the prose around them did not resolve.
        """
        score = result.resolved_share * 0.7
        signals = sum(bool(value) for value in (
            result.quantity, result.amount, result.dimensions, result.references,
            result.model_codes, result.dates, result.emails, result.lead_time_days))
        score += min(0.25, signals * 0.05)
        if result.language:
            score += 0.05 * result.language_confidence
        return round(min(1.0, score), 4)


# ===========================================================================
# Language model client
# ===========================================================================

@dataclass
class TokenUsage:
    """Running total of language-model consumption."""

    requests: int = 0
    failed_requests: int = 0
    cache_hits: int = 0
    input_tokens: int = 0
    cached_input_tokens: int = 0
    output_tokens: int = 0
    reasoning_tokens: int = 0

    # Per-million-token prices, copied from the resolved model configuration so
    # that the running total can be valued without reaching back into it.
    input_cost_per_mtok: float = INPUT_COST_PER_MTOK
    output_cost_per_mtok: float = OUTPUT_COST_PER_MTOK

    @property
    def total_tokens(self) -> int:
        return self.input_tokens + self.output_tokens

    @property
    def input_cost(self) -> float:
        """Dollar value of the input tokens consumed so far.

        Cached input is charged here at the full rate even though the provider
        discounts it heavily, so the figure leans high. That is the right
        direction for a number whose job is to stop a run becoming expensive.
        """
        return self.input_tokens / 1_000_000.0 * self.input_cost_per_mtok

    @property
    def output_cost(self) -> float:
        """Dollar value of the output tokens, reasoning tokens included."""
        return self.output_tokens / 1_000_000.0 * self.output_cost_per_mtok

    @property
    def estimated_cost(self) -> float:
        return self.input_cost + self.output_cost

    def record(self, usage: Dict[str, Any]) -> None:
        self.requests += 1
        if not usage:
            return
        self.input_tokens += int(usage.get("prompt_tokens") or usage.get("input_tokens") or 0)
        self.output_tokens += int(usage.get("completion_tokens") or usage.get("output_tokens") or 0)
        prompt_details = usage.get("prompt_tokens_details") or usage.get("input_tokens_details") or {}
        self.cached_input_tokens += int(prompt_details.get("cached_tokens") or 0)
        completion_details = (usage.get("completion_tokens_details")
                              or usage.get("output_tokens_details") or {})
        self.reasoning_tokens += int(completion_details.get("reasoning_tokens") or 0)

    def as_dict(self) -> Dict[str, Any]:
        return {
            "requests": self.requests, "failed_requests": self.failed_requests,
            "cache_hits": self.cache_hits, "input_tokens": self.input_tokens,
            "cached_input_tokens": self.cached_input_tokens,
            "output_tokens": self.output_tokens, "reasoning_tokens": self.reasoning_tokens,
            "total_tokens": self.total_tokens,
            "input_cost_per_mtok": self.input_cost_per_mtok,
            "output_cost_per_mtok": self.output_cost_per_mtok,
            "input_cost_usd": round(self.input_cost, 4),
            "output_cost_usd": round(self.output_cost, 4),
            "estimated_cost_usd": round(self.estimated_cost, 4),
        }


class SpendGuard:
    """Holds language-model spend to what the operator authorised.

    The operator names a figure when the model tier is switched on. After every
    billed response the running estimate is compared against it, and once the
    figure is reached the run pauses and asks whether to carry on. Agreeing
    raises the ceiling by the same amount again, so a limit of $25 becomes $50,
    then $75, and each further step needs its own answer. Declining switches the
    model off for the rest of the run: nothing is lost, because the local reader
    already produced an answer for every row, and the work done so far is kept.
    """

    def __init__(self, usage: TokenUsage, config: ModelConfig, interactive: bool) -> None:
        self.usage = usage
        self.config = config
        self.interactive = interactive
        self.step = max(0.0, config.spend_limit)
        self.limit = self.step
        self.extensions = 0
        self.declined = False

    @property
    def active(self) -> bool:
        """True when a limit is in force and has not been waived."""
        return bool(self.step) and not self.declined

    def as_dict(self) -> Dict[str, Any]:
        return {
            "spend_limit_usd": round(self.step, 4),
            "spend_limit_final_usd": round(self.limit, 4),
            "spend_limit_extensions": self.extensions,
            "spend_limit_stopped": self.declined,
        }

    def review(self) -> None:
        """Check the running total after a billed response and act on it."""
        if not self.active or not self.config.enabled:
            return

        # A loop rather than a single test: one costly response can overshoot
        # by more than a whole step, and each step still needs its own answer.
        while self.usage.estimated_cost >= self.limit:
            if not self.interactive:
                LOGGER.warning(
                    "Estimated language-model spend is $%.2f, at or above the $%.2f limit. "
                    "Continuing without the model; raise --llm-spend-limit to allow more.",
                    self.usage.estimated_cost, self.limit)
                self._stop()
                return
            if not self._ask():
                self._stop()
                return
            self.limit += self.step
            self.extensions += 1
            print(f"  Continuing. The limit is now ${self.limit:,.2f}.\n", flush=True)

    def _ask(self) -> bool:
        """Report the position and ask whether to keep using the model."""
        print(flush=True)
        print("=" * 79)
        print("  Language-model spend alert")
        print("=" * 79)
        print(f"  Estimated spend      : ${self.usage.estimated_cost:,.2f}")
        print(f"  Authorised so far    : ${self.limit:,.2f}")
        print(f"  Input tokens         : {self.usage.input_tokens:,} "
              f"at ${self.usage.input_cost_per_mtok:,.2f}/M = ${self.usage.input_cost:,.2f}")
        print(f"  Output tokens        : {self.usage.output_tokens:,} "
              f"at ${self.usage.output_cost_per_mtok:,.2f}/M = ${self.usage.output_cost:,.2f}")
        print(f"  Requests sent        : {self.usage.requests:,}")
        print()
        print(f"  Answering yes raises the limit to ${self.limit + self.step:,.2f}.")
        print("  Answering no finishes the run on the local reader alone.")
        try:
            answer = input("  Continue using the language model? [y/N]: ").strip().lower()
        except EOFError:
            # Input is piped and nobody is watching; the cautious reading of
            # silence is that no further spend was authorised.
            print()
            return False
        return answer[:1] == "y"

    def _stop(self) -> None:
        """Switch the model off for the remainder of the run."""
        self.declined = True
        self.config.enabled = False
        LOGGER.info(
            "Language model disabled after an estimated $%.2f of spend. "
            "The run continues on the local reader; cached answers are still used.",
            self.usage.estimated_cost)


class LanguageModelClient:
    """Minimal OpenAI-compatible chat client with disk caching."""

    def __init__(self, config: ModelConfig, cache_path: Path,
                 interactive: bool = False) -> None:
        self.config = config
        self.usage = TokenUsage(
            input_cost_per_mtok=config.input_cost_per_mtok,
            output_cost_per_mtok=config.output_cost_per_mtok,
        )
        self.guard = SpendGuard(self.usage, config, interactive)
        self.cache_path = cache_path
        self._cache: Dict[str, str] = self._load_cache()
        self._cache_dirty = False
        self._omit_temperature = False
        self._reasoning_style = "effort"

    def _load_cache(self) -> Dict[str, str]:
        if not self.cache_path.is_file():
            return {}
        try:
            payload = json.loads(self.cache_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            LOGGER.warning("Model cache %s unreadable; starting empty.", self.cache_path)
            return {}
        entries = payload.get("entries", {})
        LOGGER.info("Model cache loaded with %d entries.", len(entries))
        return entries

    def save_cache(self) -> None:
        if not self._cache_dirty:
            return
        self.cache_path.parent.mkdir(parents=True, exist_ok=True)
        payload = {
            "agent": AGENT_NAME,
            "model": self.config.model,
            "backend": self.config.backend,
            "entries": dict(sorted(self._cache.items())),
        }
        self.cache_path.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")
        self._cache_dirty = False

    def cache_key(self, task: str, payload: str) -> str:
        """Content address for one unit of work.

        The model name is part of the key so that switching backend does not
        silently reuse answers produced by a different model.
        """
        return stable_hash(task, self.config.model, payload)

    def cached(self, key: str) -> Optional[str]:
        value = self._cache.get(key)
        if value is not None:
            self.usage.cache_hits += 1
        return value

    def store(self, key: str, value: str) -> None:
        self._cache[key] = value
        self._cache_dirty = True

    def _post(self, body: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        """Issue one request, returning the decoded response or None."""
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {self.config.api_key}",
            # The shared service authenticates by api-key rather than by bearer
            # token. Sending both is accepted by each and saves branching.
            "api-key": self.config.api_key,
        }
        payload = json.dumps(body).encode("utf-8")

        try:
            if _requests is not None:
                response = _requests.post(self.config.endpoint, headers=headers,
                                          data=payload, timeout=self.config.timeout)
                status, text = response.status_code, response.text
            else:
                import urllib.error
                import urllib.request
                request = urllib.request.Request(self.config.endpoint, data=payload,
                                                 headers=headers, method="POST")
                try:
                    with urllib.request.urlopen(request, timeout=self.config.timeout) as handle:
                        status, text = handle.status, handle.read().decode("utf-8", "replace")
                except urllib.error.HTTPError as error:
                    status, text = error.code, error.read().decode("utf-8", "replace")
        except Exception as error:
            LOGGER.warning("Language-model request failed: %s", error)
            self.usage.failed_requests += 1
            return None

        if status != 200:
            summary = text[:300].replace("\n", " ")
            retry = retry_chat_body(status, text, body)
            if retry is not None:
                if "temperature" in body and "temperature" not in retry:
                    self._omit_temperature = True
                if "reasoning_effort" in body and "reasoning_effort" not in retry:
                    self._reasoning_style = "nested" if "reasoning" in retry else "omit"
                elif "reasoning" in body and "reasoning" not in retry:
                    self._reasoning_style = "omit"
                return self._post(retry)
            LOGGER.warning("Language-model request returned HTTP %s: %s", status, summary)
            self.usage.failed_requests += 1
            return None

        try:
            return json.loads(text)
        except json.JSONDecodeError:
            LOGGER.warning("Language-model response was not valid JSON.")
            self.usage.failed_requests += 1
            return None

    def complete_json(self, system_prompt: str, user_prompt: str) -> Optional[Dict[str, Any]]:
        """Request a JSON object; None on any failure."""
        if not self.config.enabled:
            return None
        if self.config.max_requests and self.usage.requests >= self.config.max_requests:
            LOGGER.warning("Language-model request cap (%d) reached.", self.config.max_requests)
            return None

        body: Dict[str, Any] = chat_completion_body(
            self.config.model, system_prompt, user_prompt,
            omit_temperature=self._omit_temperature,
            reasoning_effort=getattr(self.config, "reasoning_effort", DEFAULT_REASONING_EFFORT),
            reasoning_style=self._reasoning_style,
        )

        response = self._post(body)
        if response is None:
            return None
        self.usage.record(response.get("usage") or {})
        # Checked after every response, so a limit reached mid-batch takes
        # effect on the next call rather than at the end of the phase.
        self.guard.review()
        try:
            content = response["choices"][0]["message"]["content"]
        except (KeyError, IndexError, TypeError):
            return None
        return extract_json_object(content or "")


def extract_json_object(content: str) -> Optional[Dict[str, Any]]:
    """Recover a JSON object from a model reply.

    Even with a JSON response format enforced, replies occasionally arrive
    wrapped in a code fence or with a sentence in front of them, so the object
    is located by brace balance rather than assumed to be the whole string.
    """
    content = content.strip()
    if content.startswith("```"):
        content = re.sub(r"^```[a-zA-Z]*\s*|\s*```$", "", content).strip()
    try:
        parsed = json.loads(content)
        return parsed if isinstance(parsed, dict) else None
    except json.JSONDecodeError:
        pass

    start = content.find("{")
    if start < 0:
        return None
    depth, in_string, escape = 0, False, False
    for index in range(start, len(content)):
        char = content[index]
        if in_string:
            if escape:
                escape = False
            elif char == "\\":
                escape = True
            elif char == '"':
                in_string = False
            continue
        if char == '"':
            in_string = True
        elif char == "{":
            depth += 1
        elif char == "}":
            depth -= 1
            if depth == 0:
                try:
                    parsed = json.loads(content[start:index + 1])
                    return parsed if isinstance(parsed, dict) else None
                except json.JSONDecodeError:
                    return None
    return None


_MODEL_CACHE_PREFIX = "read_en_v2"

_MODEL_SYSTEM_PROMPT = (
    "You read short procurement free-text lines written in Finnish, Swedish, "
    "Polish, German or English and describe what was bought.\n"
    "EVERY value you return MUST be English. Translate every Finnish, Swedish, "
    "Polish or German common noun. Never copy a source-language word into any "
    "field. Keep part numbers, brand names, place names and person names as they "
    "are; translate everything else.\n"
    "Return JSON only, in the form {\"items\": [{\"id\": \"...\", "
    "\"description\": \"...\", \"type\": \"Material|Service\", "
    "\"intent\": \"...\"}]}.\n"
    "Rules:\n"
    "- description: plain English noun phrase, at most twelve words, naming the "
    "goods or the activity. No invoice numbers, no supplier names, no prices, "
    "no source-language words.\n"
    "- Use the 'text' field as the authoritative purchase line. The optional "
    "'context' field may come from a different document after a join; ignore "
    "it when it does not describe the same purchase.\n"
    "- If 'text' is only a category or material group, keep that category in "
    "English and do not invent a specific item.\n"
    "- type: Material for a physical item, Service for an activity.\n"
    "- intent: a short English category such as Maintenance and repair, Spare "
    "parts and materials, Consulting and advisory, Travel and accommodation, "
    "Freight and delivery, Rental and leasing, Training and certification, "
    "Licences and subscriptions, Fees and charges. Use an empty string when "
    "unsure. Intent must be English.\n"
    "- Return one entry for every id supplied, in the same order. Never invent "
    "detail that is not present in the text."
)

_MODEL_REVIEW_PROMPT = (
    "You review procurement interpretations for two faults: leftover "
    "non-English words, and a description that does not match the purchase "
    "line.\n"
    "EVERY value you return MUST be English. Translate leftover Finnish, "
    "Swedish, Polish or German common nouns. Keep part numbers, brands, place "
    "names and person names.\n"
    "Return JSON only, in the form {\"items\": [{\"id\": \"...\", "
    "\"description\": \"...\", \"type\": \"Material|Service\", "
    "\"intent\": \"...\"}]}.\n"
    "Rules:\n"
    "- 'text' is the authoritative purchase line. Ignore 'context' when it "
    "describes a different purchase (for example an invoice article in another "
    "language that does not match the line).\n"
    "- If the draft description is irrelevant to 'text', rewrite it from "
    "'text' only.\n"
    "- If 'text' is empty or a placeholder, use a cautious English reading of "
    "whatever category remains; do not invent a product.\n"
    "- description: English noun phrase, at most twelve words.\n"
    "- type and intent: English, same allowed values as the reader prompt.\n"
    "- Return one entry for every id, in the same order. Never invent detail."
)


class ModelReader:
    """Asks the model to read the free text the local reader could not.

    Works on distinct strings rather than rows, in batches, with every answer
    cached on a hash of the text. A second run over the same data therefore
    costs nothing, and the answers cannot drift between runs.
    """

    def __init__(self, client: LanguageModelClient, settings: Settings) -> None:
        self.client = client
        self.settings = settings
        self.resolved = 0

    def resolve(self, texts: Sequence[str],
                contexts: Optional[Dict[str, str]] = None,
                review: bool = False) -> Dict[str, Dict[str, str]]:
        """Read a list of distinct strings, returning text -> fields.

        ``contexts`` is optional extra evidence, keyed by the same string as
        ``texts``. It is sent to the model but is not part of the cache key:
        the purchase line is what was bought, and a joined invoice article
        must not produce a different description of the same line.
        """
        answers: Dict[str, Dict[str, str]] = {}
        pending: List[str] = []
        contexts = contexts or {}
        prefix = "review_en_v1" if review else _MODEL_CACHE_PREFIX

        for text in texts:
            cached = self.client.cached(self.client.cache_key(prefix, text))
            if cached is not None:
                try:
                    answers[text] = json.loads(cached)
                except json.JSONDecodeError:
                    pending.append(text)
            else:
                pending.append(text)

        if not pending:
            return answers

        batch_size = max(1, self.settings.model.batch_size)
        # Sorted so that the batching, and therefore the cache keys of any
        # future run over the same data, are identical every time.
        pending.sort()
        prompt = _MODEL_REVIEW_PROMPT if review else _MODEL_SYSTEM_PROMPT
        for start in range(0, len(pending), batch_size):
            batch = pending[start:start + batch_size]
            if not self.client.config.enabled:
                break
            resolved = self._ask(batch, contexts, prompt)
            for text, fields in resolved.items():
                answers[text] = fields
                self.client.store(self.client.cache_key(prefix, text),
                                  json.dumps(fields, ensure_ascii=False, sort_keys=True))
            self.resolved += len(resolved)
            LOGGER.info("Model %s %d of %d descriptions.",
                        "reviewed" if review else "read",
                        min(start + batch_size, len(pending)), len(pending))
        return answers

    def _ask(self, batch: Sequence[str], contexts: Dict[str, str],
             prompt: str) -> Dict[str, Dict[str, str]]:
        identifiers = {stable_hash(text): text for text in batch}
        payload = {"items": [
            {"id": key, "text": value, "context": contexts.get(value, "")}
            for key, value in sorted(identifiers.items())
        ]}
        response = self.client.complete_json(prompt, json.dumps(payload, ensure_ascii=False))
        if not response:
            return {}

        results: Dict[str, Dict[str, str]] = {}
        for item in response.get("items") or []:
            if not isinstance(item, dict):
                continue
            text = identifiers.get(normalise_text(item.get("id")))
            if text is None:
                continue
            description = drop_foreign_common_nouns(normalise_text(item.get("description")))
            if not description:
                continue
            item_type = normalise_text(item.get("type")).title()
            intent = drop_foreign_common_nouns(normalise_text(item.get("intent")))
            results[text] = {
                "description": description,
                "type": item_type if item_type in {"Material", "Service"} else "",
                "intent": intent,
            }
        return results


# ===========================================================================
# Stage 1: invoice lines onto Sievo
# ===========================================================================

@dataclass
class InvoiceLine:
    """One invoice row, reduced to the fields the join and the output need."""

    values: Dict[str, str]
    is_line: bool
    row_total_excl: Optional[float]
    row_total_incl: Optional[float]
    vat_amount: Optional[float]


class InvoiceIndex:
    """Every invoice line, indexed by each key the join is willing to try.

    Duplicates are dropped on the way in. An invoice cannot legitimately carry
    the same line number twice, so a repeated pair is an artefact of the export
    and summing it would overstate the invoice.
    """

    # Ordered most specific first. Each entry is (name, is_line_level).
    STRATEGIES: Tuple[Tuple[str, bool], ...] = (
        ("invoice-number + line", True),
        ("document-number + line", True),
        ("document-id + line", True),
        ("invoice-number", False),
        ("document-number", False),
        ("document-id", False),
    )

    def __init__(self, tables: Sequence[Table]) -> None:
        self.lines: List[InvoiceLine] = []
        self.columns: Dict[str, str] = {}
        self.duplicates = 0
        self.summary_rows = 0

        self._by_line: Dict[Tuple[str, str], List[int]] = defaultdict(list)
        self._by_document: Dict[str, List[int]] = defaultdict(list)
        self._document_totals: Dict[str, Tuple[int, float]] = {}

        seen: Set[Tuple[str, str, str]] = set()
        for table in tables:
            mapping = ColumnMap(table.headers, INVOICE_FIELDS)
            self.columns.update(mapping.resolved)
            for _, record in table.iter_records():
                self._ingest(mapping, record, seen)

        self._summarise()

    def _ingest(self, mapping: ColumnMap, record: Dict[str, str],
                seen: Set[Tuple[str, str, str]]) -> None:
        values = {field_name: mapping.get(record, field_name) for field_name in INVOICE_FIELDS}
        key = values.get("invoice_key") or values.get("invoice_id")
        if is_blank(key) and is_blank(values.get("xml_file_name")):
            return

        row = line_key(values.get("row_number"))
        signature = (compact_key(key), row, lookup_key(values.get("article_name", "")))
        if signature in seen:
            self.duplicates += 1
            return
        seen.add(signature)

        # A row without a line number is a header or a total, not a purchase
        # line; including it in a sum would double-count the invoice.
        is_line = bool(row) and row.isdigit()
        if not is_line:
            self.summary_rows += 1

        line = InvoiceLine(
            values=values,
            is_line=is_line,
            row_total_excl=parse_amount(values.get("row_total_excl_vat")),
            row_total_incl=parse_amount(values.get("row_total_incl_vat")),
            vat_amount=parse_amount(values.get("vat_amount")),
        )
        index = len(self.lines)
        self.lines.append(line)

        for document in self._document_keys(values):
            self._by_document[document].append(index)
            if row:
                self._by_line[(document, row)].append(index)

    @staticmethod
    def _document_keys(values: Dict[str, str]) -> Set[str]:
        """Every key under which this invoice line can be found."""
        keys: Set[str] = set()
        for field_name in ("invoice_key", "invoice_id"):
            value = values.get(field_name, "")
            if not is_blank(value):
                keys.add(compact_key(value))
        document = document_id_key(values.get("xml_file_name", ""))
        if document:
            keys.add(document)
        return keys

    def _summarise(self) -> None:
        """Pre-compute the line count and total for each invoice."""
        for document, indices in self._by_document.items():
            lines = [self.lines[index] for index in indices if self.lines[index].is_line]
            total = sum(line.row_total_excl or 0.0 for line in lines)
            self._document_totals[document] = (len(lines), total)

    @property
    def empty(self) -> bool:
        return not self.lines

    def lookup(self, sievo: Dict[str, str]) -> Tuple[str, List[InvoiceLine], str]:
        """Find the invoice lines for one Sievo row.

        Returns (strategy, lines, document_key). The ladder is walked in order
        and the first strategy that produces anything wins, so a line-level
        match is never displaced by a looser one.
        """
        invoice_number = "" if is_blank(sievo.get("invoice_number")) else compact_key(sievo["invoice_number"])
        document_number = "" if is_blank(sievo.get("document_number")) else compact_key(sievo["document_number"])
        document_id = document_id_key(sievo.get("document_identifier") or "") \
            or document_id_key(sievo.get("invoice_link") or "")
        row = line_key(sievo.get("document_line"))

        candidates = (
            ("invoice-number + line", invoice_number, True),
            ("document-number + line", document_number, True),
            ("document-id + line", document_id, True),
            ("invoice-number", invoice_number, False),
            ("document-number", document_number, False),
            ("document-id", document_id, False),
        )

        for name, key, line_level in candidates:
            if not key:
                continue
            if line_level:
                if not row:
                    continue
                indices = self._by_line.get((key, row))
            else:
                indices = self._by_document.get(key)
            if indices:
                return name, [self.lines[index] for index in indices], key
        return "", [], ""

    def document_summary(self, document: str) -> Tuple[int, float]:
        return self._document_totals.get(document, (0, 0.0))


# The invoice block appended to every row in stage 1.
INVOICE_OUTPUT_FIELDS: Tuple[str, ...] = (
    "invoice_key", "invoice_id", "row_number", "xml_file_name", "article_id",
    "article_name", "free_text", "quantity_charged", "quantity_delivered",
    "unit_price_excl_vat", "unit_price_net", "row_total_excl_vat",
    "row_total_incl_vat", "vat_amount", "vat_rate",
)

INVOICE_COLUMNS: List[str] = (
    ["Invoice_Match_Level", "Invoice_Match_Strategy", "Invoice_Match_Count"]
    + [_column("Invoice", name) for name in INVOICE_OUTPUT_FIELDS]
    + ["Invoice_Lines_On_Document", "Invoice_Document_Total_Excl_Vat",
       "Invoice_Matched_Total_Excl_Vat", "Invoice_Article_Names_All"]
)


def join_invoice(sievo: Dict[str, str], index: Optional[InvoiceIndex]) -> Dict[str, str]:
    """Produce the invoice block for one Sievo row.

    One row in, one block out. Where several invoice lines answer to the
    transaction, scalar columns are filled only where every candidate agrees,
    and the rest of the evidence is summarised.
    """
    block = {column: "" for column in INVOICE_COLUMNS}
    block["Invoice_Match_Level"] = "none"
    if index is None or index.empty:
        return block

    strategy, lines, document = index.lookup(sievo)
    if not lines:
        return block

    line_rows = [line for line in lines if line.is_line] or lines
    block["Invoice_Match_Strategy"] = strategy
    block["Invoice_Match_Count"] = str(len(line_rows))
    block["Invoice_Match_Level"] = "line" if strategy.endswith("+ line") else "header"

    if len(line_rows) == 1:
        for field_name in INVOICE_OUTPUT_FIELDS:
            block[_column("Invoice", field_name)] = line_rows[0].values.get(field_name, "")
    else:
        # Fill only where the candidates agree; a disagreement is left blank
        # rather than resolved arbitrarily.
        for field_name in INVOICE_OUTPUT_FIELDS:
            distinct = {line.values.get(field_name, "") for line in line_rows}
            distinct.discard("")
            if len(distinct) == 1:
                block[_column("Invoice", field_name)] = distinct.pop()

    names = sorted({line.values.get("article_name", "") for line in line_rows} - {""})
    block["Invoice_Article_Names_All"] = "; ".join(names[:12])
    block["Invoice_Matched_Total_Excl_Vat"] = format_amount(
        sum(line.row_total_excl or 0.0 for line in line_rows) or None)

    count, total = index.document_summary(document)
    block["Invoice_Lines_On_Document"] = str(count) if count else ""
    block["Invoice_Document_Total_Excl_Vat"] = format_amount(total or None)
    return block


# ===========================================================================
# Stage 2: purchase order lines onto the widened table
# ===========================================================================

class PurchaseOrderIndex:
    """Maximo and Basware lines, indexed by PO number and PO line number."""

    def __init__(self, system: str, tables: Sequence[Table],
                 synonyms: Dict[str, Sequence[str]]) -> None:
        self.system = system
        self.native_headers: List[str] = []
        self.records: List[Dict[str, str]] = []
        self.mapped: List[Dict[str, str]] = []

        self._by_line: Dict[Tuple[str, str], List[int]] = defaultdict(list)
        self._by_order: Dict[str, List[int]] = defaultdict(list)

        seen_headers: Set[str] = set()
        for table in tables:
            mapping = ColumnMap(table.headers, synonyms)
            for header in table.headers:
                if header and header not in seen_headers:
                    seen_headers.add(header)
                    self.native_headers.append(header)

            for _, record in table.iter_records():
                values = {name: mapping.get(record, name) for name in synonyms}
                order = values.get("po_number", "")
                if is_blank(order):
                    continue
                order_key = compact_key(order)
                index = len(self.records)
                self.records.append({header: normalise_text(record.get(header, ""))
                                     for header in table.headers})
                self.mapped.append(values)
                self._by_order[order_key].append(index)
                row = line_key(values.get("po_line_number"))
                if row:
                    self._by_line[(order_key, row)].append(index)

    @property
    def empty(self) -> bool:
        return not self.records

    def lookup(self, order: str, row: str) -> Tuple[str, List[int]]:
        """Find the PO lines for one transaction.

        A missing line number is common in the transaction extract. When the
        order has exactly one line the match is still unambiguous and is taken;
        when it has several, only the header-level fields can be trusted.
        """
        order_key = compact_key(order)
        if not order_key or order_key not in self._by_order:
            return "", []

        if row:
            indices = self._by_line.get((order_key, row))
            if indices:
                return "po-number + line", indices

        indices = self._by_order[order_key]
        if len(indices) == 1:
            return "po-number, single-line order", indices
        return "po-number", indices

    def order_line_count(self, order: str) -> int:
        return len(self._by_order.get(compact_key(order), ()))


PO_CORE_COLUMNS: List[str] = (
    ["PO_Match_Level", "PO_Match_Strategy", "PO_Match_Count", "PO_System",
     "PO_Lines_On_Order"]
    + [_column("PO", name) for name in PO_OUTPUT_FIELDS]
)


class PurchaseOrderJoiner:
    """Attaches the PO line to a transaction, from whichever system holds it."""

    def __init__(self, indexes: Dict[str, PurchaseOrderIndex], native: bool) -> None:
        self.indexes = {name: index for name, index in indexes.items() if not index.empty}
        self.native = native
        self.native_columns: List[str] = []
        self._native_index: Dict[str, List[str]] = {}
        if native:
            for name, index in sorted(self.indexes.items()):
                prefix = name.capitalize()
                columns = [f"{prefix}_{header}" for header in index.native_headers]
                self._native_index[name] = columns
                self.native_columns.extend(columns)

    @property
    def columns(self) -> List[str]:
        return PO_CORE_COLUMNS + self.native_columns

    def join(self, sievo: Dict[str, str]) -> Dict[str, str]:
        block = {column: "" for column in self.columns}
        block["PO_Match_Level"] = "none"
        if not self.indexes:
            return block

        order = sievo.get("po_number", "")
        if is_blank(order):
            return block
        row = line_key(sievo.get("po_line"))

        best = self._select(order, row, sievo.get("data_source", ""))
        if best is None:
            return block

        system, strategy, indices = best
        index = self.indexes[system]
        block["PO_System"] = system.capitalize()
        block["PO_Match_Strategy"] = strategy
        block["PO_Match_Count"] = str(len(indices))
        block["PO_Match_Level"] = "line" if len(indices) == 1 else "header"
        block["PO_Lines_On_Order"] = str(index.order_line_count(order))

        if len(indices) == 1:
            values = index.mapped[indices[0]]
            for name in PO_OUTPUT_FIELDS:
                block[_column("PO", name)] = values.get(name, "")
        else:
            # Only unanimous values are trustworthy at header level.
            for name in PO_OUTPUT_FIELDS:
                distinct = {index.mapped[position].get(name, "") for position in indices}
                distinct.discard("")
                if len(distinct) == 1:
                    block[_column("PO", name)] = distinct.pop()

        if self.native:
            self._fill_native(block, system, index, indices)
        return block

    def _select(self, order: str, row: str,
                data_source: str) -> Optional[Tuple[str, str, List[int]]]:
        """Choose which PO system answers for this transaction.

        The transaction extract usually names its own source system, and that
        naming is trusted when it resolves. Otherwise both systems are probed
        and the more specific match wins; a tie is broken by system name so the
        result does not depend on dictionary order.
        """
        hint = lookup_key(data_source).replace("x.", "").strip()
        preferred = [name for name in self.indexes if name in hint]

        found: List[Tuple[int, str, str, List[int]]] = []
        for name in sorted(self.indexes):
            strategy, indices = self.indexes[name].lookup(order, row)
            if not indices:
                continue
            rank = 0 if strategy.startswith("po-number + line") else (
                1 if "single-line" in strategy else 2)
            if name in preferred:
                rank -= 1
            found.append((rank, name, strategy, indices))

        if not found:
            return None
        found.sort(key=lambda item: (item[0], item[1]))
        _, name, strategy, indices = found[0]
        return name, strategy, indices

    def _fill_native(self, block: Dict[str, str], system: str,
                     index: PurchaseOrderIndex, indices: List[int]) -> None:
        columns = self._native_index.get(system) or []
        headers = index.native_headers
        if len(indices) == 1:
            record = index.records[indices[0]]
            for column, header in zip(columns, headers):
                block[column] = record.get(header, "")
            return
        for column, header in zip(columns, headers):
            distinct = {index.records[position].get(header, "") for position in indices}
            distinct.discard("")
            if len(distinct) == 1:
                block[column] = distinct.pop()


# ===========================================================================
# Orchestration
# ===========================================================================

# The free-text fields read in stage 3, in the order they are consulted.
#
# Primary fields say what was bought and are the only ones that shape the
# generated description when they are populated. Fallback fields — a joined
# invoice article, a material group — are used only when the line itself is
# silent. Concatenating a joined invoice onto a filled line is what used to
# turn a UK "Booking Fee" into "Booking fee kuljetukset" after a false join.
# Supporting fields (notes, invoice free text, item codes) are scanned for
# dates, references and quantities, then set aside.
TEXT_SOURCE_COLUMNS: Tuple[Tuple[str, str, str], ...] = (
    ("Document line desc", "sievo.document-line", "primary"),
    ("PO line desc", "sievo.po-line", "primary"),
    ("PO_Line_Description", "po.line", "primary"),
    ("PO_Header_Description", "po.header", "primary"),
    ("Invoice_Article_Name", "invoice.article", "fallback"),
    ("Invoice_Article_Names_All", "invoice.articles", "fallback"),
    ("MaterialGroupName", "sievo.material-group", "fallback"),
    ("PO_Item_Code", "po.item", "supporting"),
    ("PO_Internal_Note", "po.note", "supporting"),
    ("Invoice_Free_Text", "invoice.free-text", "supporting"),
)


# ===========================================================================
# Stages 4 to 6: the agents
# ===========================================================================

def read_csv_rows(path: Path) -> Iterator[Dict[str, str]]:
    """Stream a stage file back off disk, one row at a time.

    Streaming rather than loading: the stages are written to be read by the next
    stage, and a million-row wide table does not want to be in memory twice.
    """
    with path.open("r", encoding=CSV_ENCODING, newline="") as handle:
        for record in csv.DictReader(handle):
            yield {key: (value or "") for key, value in record.items() if key is not None}


def _header_of(path: Path) -> List[str]:
    """Column names of a CSV, or an empty list if it cannot be read."""
    try:
        with path.open("r", encoding=CSV_ENCODING, newline="") as handle:
            for row in csv.reader(handle):
                return [name.strip() for name in row]
    except OSError:
        return []
    return []


def _now() -> str:
    """A timestamp a reader can compare against a file's modification time."""
    return datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _duration(seconds: float) -> str:
    """A duration in the units a reader thinks in."""
    if seconds < 90:
        return f"{seconds:.0f}s"
    minutes, rest = divmod(int(seconds), 60)
    if minutes < 90:
        return f"{minutes}m {rest:02d}s"
    hours, minutes = divmod(minutes, 60)
    return f"{hours}h {minutes:02d}m"


def _digest(path: Path) -> str:
    """Content digest of a file, or an empty string if it cannot be read."""
    hasher = hashlib.sha256()
    try:
        with path.open("rb") as handle:
            for block in iter(lambda: handle.read(1 << 20), b""):
                hasher.update(block)
    except OSError:
        return ""
    return hasher.hexdigest()[:16]


@dataclass
class AgentRun:
    """What one agent produced, and whether it can be merged."""

    name: str
    script: str
    output: Path
    introduced: Tuple[str, ...] = ()
    ok: bool = False
    reason: str = ""
    rows: int = 0

    # Short tag prefixing this agent's log lines as they are forwarded.
    tag: str = ""

    reused: bool = False
    seconds: float = 0.0
    log: Optional[Path] = None


class AgentPipeline:
    """Runs Agents 1 to 4 over the stage-3 file and widens it with their columns.

    The agents are run as separate processes rather than imported. Each is a
    standalone script carrying its own copy of the vocabulary loader, its own
    response cache and its own spend guard, and each writes its own manifest that
    the client reads. Keeping them in their own process is what lets Max add their
    columns without taking on responsibility for their internals, and means a
    change inside an agent cannot break the builder except at the one place the
    two meet - the column names asserted above.

    The governing rule of the earlier stages continues to apply: a stage widens
    the table and never lengthens it. Agents 1 and 3 annotate a row at a time and
    Agent 2 only appends, so a row count that changes here is a defect and stops
    the run rather than being reported as a result.

    Agent 4 is the exception and is not merged at all. Its unit of analysis is a
    supplier within a comparison scope, not a purchase line, so there is no
    honest one-to-one mapping onto a transaction row. It is written alongside as a
    companion file, joinable on the supplier key that this class adds to the wide
    row.
    """

    def __init__(self, settings: Settings) -> None:
        self.settings = settings
        self.here = Path(__file__).resolve().parent
        self.work = settings.results_dir / "agents"
        self.outputs: List[str] = []
        self.statistics: Dict[str, Any] = {}
        self.runs: List[AgentRun] = []
        self.supplier_keys: Dict[str, str] = {}

        # Agents currently running, so an interrupt can end them rather than
        # leaving one working on behalf of a run that has stopped.
        self._running: List[subprocess.Popen] = []

    # -- running the agents --------------------------------------------------

    def _common_arguments(self, results: Path) -> List[str]:
        """Arguments every agent takes, so one Max run configures all four."""
        arguments = ["--non-interactive",
                     "--results", str(results),
                     "--lexicon", str(self.settings.lexicon_path),
                     "--cache", str(self.settings.cache_dir)]
        if not self.settings.write_jsonl:
            arguments.append("--no-jsonl")
        if self.settings.model.enabled:
            arguments.append("--use-llm")
            if self.settings.model.spend_limit:
                arguments += ["--llm-spend-limit", f"{self.settings.model.spend_limit:.2f}"]
        return arguments

    def _invoke(self, run: AgentRun, arguments: Sequence[str],
                input_csv: Optional[Path] = None) -> AgentRun:
        """Reuse or run one agent, reporting what happened without raising."""
        script = self.here / run.script
        if not script.is_file():
            run.reason = f"{run.script} is not in {self.here}"
            LOGGER.error("Cannot run %s: %s.", run.name, run.reason)
            return run

        if input_csv is not None:
            cached = self._reusable(run, input_csv, arguments)
            if cached:
                run.ok = True
                run.reused = True
                run.reason = "already produced from this input by this script"
                self._say(run, f"reusing {run.output.name} - {run.reason}")
                return run

        returncode, tail = self._stream(run, arguments)

        if returncode is None:
            run.reason = (f"said nothing for {self.settings.agent_silence_timeout}s "
                          f"and was treated as hung")
            self._say(run, f"NOT RESPONDING - {run.reason}")
            return run
        if returncode == "over":
            run.reason = f"ran past the {self.settings.agent_timeout}s ceiling"
            self._say(run, f"STOPPED - {run.reason}")
            return run

        if returncode != 0:
            # The agent's own last words are far more useful than the exit code.
            run.reason = tail[-1] if tail else f"exit code {returncode}"
            self._say(run, f"FAILED: {run.reason}")
            return run

        if not run.output.is_file():
            run.reason = f"finished but wrote no {run.output.name}"
            self._say(run, run.reason)
            return run

        run.ok = True
        self._record(run, input_csv, arguments)
        self._say(run, f"done in {_duration(run.seconds)}")
        return run

    # -- not doing the same work twice ---------------------------------------

    def _fingerprint(self, run: AgentRun, input_csv: Path,
                     arguments: Sequence[str]) -> Dict[str, Any]:
        """What a result on disk has to match before it is worth reusing.

        The input's contents, the agent script's contents, and the settings that
        change what the agent does. Digesting the script rather than reading its
        version means editing an agent invalidates its old output whether or not
        anyone remembered to bump the number.
        """
        settings = sorted(argument for argument in arguments
                          if argument.startswith("--")
                          and argument not in {"--results", "--lexicon", "--cache",
                                               "--input", "--registry",
                                               "--catalogues", "--reference",
                                               "--non-interactive", "--no-jsonl"})
        # Paths differ harmlessly between machines, so what a path points at is
        # digested rather than the path itself. Agent 3's catalogue is the case
        # that matters: Fortum extend the master as they go, and matches found
        # against last month's copy are not answers to the question being asked.
        reference = ""
        for flag in ("--catalogues", "--reference"):
            if flag in arguments:
                target = Path(arguments[list(arguments).index(flag) + 1])
                if target.is_file():
                    reference = _digest(target)
                break
        return {"input": _digest(input_csv),
                "script": _digest(self.here / run.script),
                "lexicon": _digest(self.settings.lexicon_path),
                "reference": reference,
                "settings": settings}

    def _stamp(self, run: AgentRun) -> Path:
        return self.work / f".{Path(run.script).stem}.reuse.json"

    def _reusable(self, run: AgentRun, input_csv: Path,
                  arguments: Sequence[str]) -> bool:
        """Whether this agent's result is already on disk and still valid.

        Reuse matters most for the agent that costs the most. Agent 1 on the
        client's full extract runs for hours, and without this a build interrupted
        anywhere after it - by a failure further down the chain, or by the operator
        - pays for those hours again to reach the same file.
        """
        if self.settings.force_agents or not run.output.is_file():
            return False
        stamp = self._stamp(run)
        if not stamp.is_file():
            return False
        try:
            recorded = json.loads(stamp.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            return False
        return recorded.get("fingerprint") == self._fingerprint(run, input_csv,
                                                               arguments)

    def _record(self, run: AgentRun, input_csv: Optional[Path],
                arguments: Sequence[str]) -> None:
        """Note what this result was made from, so it can be reused."""
        if input_csv is None:
            return
        try:
            self._stamp(run).write_text(json.dumps({
                "fingerprint": self._fingerprint(run, input_csv, arguments),
                "output": run.output.name,
                "seconds": round(run.seconds, 1),
            }, indent=2), encoding="utf-8")
        except OSError as error:
            LOGGER.debug("Could not record what %s was made from: %s",
                         run.output.name, error)

    # -- watching an agent work ---------------------------------------------

    def _say(self, run: AgentRun, message: str) -> None:
        print(f"  [{run.tag or run.name}] {message}", flush=True)

    def stop_running_agents(self) -> None:
        """End any agent still running, so Ctrl-C does not leave one behind.

        A child started with Popen keeps going after the parent stops waiting for
        it, and an abandoned Agent 3 would carry on embedding a catalogue for a run
        that no longer exists - burning the machine and, worse, writing its output
        after the interrupted run had already decided it had none.
        """
        for process in list(self._running):
            if process.poll() is None:
                process.terminate()
        for process in list(self._running):
            try:
                process.wait(timeout=10)
            except subprocess.TimeoutExpired:
                process.kill()
        self._running.clear()

    def _stream(self, run: AgentRun,
                arguments: Sequence[str]) -> Tuple[Any, List[str]]:
        """Run one agent, forwarding its own logging as it happens.

        The output used to be captured and thrown away unless the agent failed, and
        then only its last line of standard error was kept. On the client's full
        extract that meant six hours during which the build said nothing at all,
        followed by a timeout whose message could not say what the agent had been
        doing or how far it had got.

        Now every line is forwarded with the agent's name in front of it and kept
        in a log file, and the time of the last line is what the watchdog judges.
        Returns the exit code, or None where the agent went quiet for too long, or
        the string ``"over"`` where it passed an absolute ceiling that was asked for.
        """
        self.work.mkdir(parents=True, exist_ok=True)
        run.log = self.work / f"{Path(run.script).stem}.log"
        command = [sys.executable, str(self.here / run.script), *arguments]

        self._say(run, "starting")
        LOGGER.info("%s: %s", run.name, " ".join(arguments[:6]))

        started = time.time()
        tail: List[str] = []

        # When the agent last said something, and when the build last said the
        # agent was alive. Two clocks, not one: a heartbeat is the build talking
        # about the agent, so counting it as the agent talking would keep the
        # silence below the limit for ever and the watchdog would never fire.
        last_line = [started]
        last_beat = started

        try:
            process = subprocess.Popen(
                command, cwd=str(self.here), text=True, bufsize=1,
                stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                # Without this a Python child writing to a pipe holds its output
                # until it exits, and a live log is not live.
                env={**os.environ, "PYTHONUNBUFFERED": "1"})
        except OSError as error:
            run.seconds = time.time() - started
            return 1, [str(error)]

        self._running.append(process)

        def pump() -> None:
            assert process.stdout is not None
            with run.log.open("w", encoding="utf-8") as log:
                for line in process.stdout:
                    text = line.rstrip("\n")
                    last_line[0] = time.time()
                    log.write(line)
                    if text.strip():
                        tail.append(text.strip())
                        del tail[:-40]
                        print(f"  [{run.tag or run.name}] {text}", flush=True)

        reader = threading.Thread(target=pump, name=f"{run.tag}-log", daemon=True)
        reader.start()

        verdict: Any = None
        while True:
            try:
                verdict = process.wait(timeout=5)
                break
            except subprocess.TimeoutExpired:
                pass

            elapsed = time.time() - started
            quiet = time.time() - last_line[0]

            if self.settings.agent_timeout and elapsed > self.settings.agent_timeout:
                verdict = "over"
                break
            if quiet > self.settings.agent_silence_timeout:
                verdict = None
                break
            # A heartbeat while the agent is quiet but not yet suspect. Agent 1
            # embeds and translates in long silent passes, and a build that prints
            # nothing for an hour gets killed by whoever is watching it.
            if quiet >= HEARTBEAT_SECONDS and time.time() - last_beat >= HEARTBEAT_SECONDS:
                last_beat = time.time()
                self._say(run, f"still working, {_duration(elapsed)} elapsed, "
                               f"quiet for {_duration(quiet)}")

        if verdict is None or verdict == "over":
            process.kill()
            process.wait()
        reader.join(timeout=10)
        if process in self._running:
            self._running.remove(process)
        run.seconds = time.time() - started
        return verdict, tail

    def _catalogue_arguments(self) -> List[str]:
        """Where Agent 3 should look for item catalogues.

        The client's catalogue master is named for what it is, so it is looked for
        by name and the newest, largest copy wins wherever it sits.

        This used to hand over the ./catalogues folder whenever that folder
        existed, without looking inside it. That silently preferred a 4,200-item
        copy left in ./catalogues to the 846,000-item master the client had since
        sent into the source folder, and Agent 3 duly matched nothing against it -
        a run that looks complete and reports far less than it should. Fortum send
        the master as a full replacement rather than as a delta, so a smaller file
        of the same name is an older copy of the same thing and never an addition
        to it.
        """
        found: List[Tuple[Path, str]] = []
        for folder, label in ((self.settings.source_dir, "source folder"),
                              (self.here / "catalogues", "catalogues folder")):
            if not folder or not folder.is_dir():
                continue
            for path in sorted(folder.glob(CATALOGUE_MASTER_GLOB)):
                if path.is_file() and not path.name.startswith((".", "~$")):
                    found.append((path, label))

        if found:
            found.sort(key=lambda entry: (-entry[0].stat().st_size, entry[0].name))
            chosen, label = found[0]
            LOGGER.info("Agent 3 catalogue: %s from the %s (%s bytes).",
                        chosen.name, label, f"{chosen.stat().st_size:,}")
            for other, where in found[1:]:
                if other.stat().st_size * 4 < chosen.stat().st_size:
                    LOGGER.warning(
                        "%s in the %s is far smaller than the master being used "
                        "(%s against %s bytes) and looks like an older copy. It has "
                        "not been read; delete it to avoid doubt.",
                        other.name, where, f"{other.stat().st_size:,}",
                        f"{chosen.stat().st_size:,}")
            return ["--catalogues", str(chosen)]

        catalogues = self.here / "catalogues"
        if catalogues.is_dir():
            LOGGER.warning("No file named like the client's catalogue master was "
                           "found, so every catalogue in %s is read instead.",
                           catalogues)
            return ["--catalogues", str(catalogues)]

        # Safe as a last resort because Agent 3 refuses any file carrying
        # purchase-transaction columns, and the source folder is full of them.
        LOGGER.warning("No item catalogue was found. Agent 3 will read %s and "
                       "refuse the purchase extracts in it, leaving it little to "
                       "match against.", self.settings.source_dir)
        return ["--reference", str(self.settings.source_dir)]

    def run_agents(self, stage3_csv: Path) -> None:
        """Run all four agents over the stage-3 master file."""
        self.work.mkdir(parents=True, exist_ok=True)
        common = self._common_arguments(self.work)

        agent1 = self._invoke(
            AgentRun("Agent 1 - purchase descriptions", "agent1.py",
                     self.work / "agent1_unified_lines.csv", AGENT1_REQUIRED,
                     tag="Agent 1"),
            [*common, "--input", str(stage3_csv)], stage3_csv)
        self.runs.append(agent1)
        if not agent1.ok:
            return

        agent2 = self._invoke(
            AgentRun("Agent 2 - purchase groups", "agent2.py",
                     self.work / "agent2_purchase_groups.csv", AGENT2_REQUIRED,
                     tag="Agent 2"),
            [*common, "--input", str(agent1.output),
             "--registry", str(self.here / "lexicon" / "agent2_group_registry.json")],
            agent1.output)
        self.runs.append(agent2)
        if not agent2.ok:
            return

        agent3 = self._invoke(
            AgentRun("Agent 3 - standard items", "agent3.py",
                     self.work / "agent3_standardisation.csv", AGENT3_REQUIRED,
                     tag="Agent 3"),
            [*common, "--input", str(agent2.output), *self._catalogue_arguments()],
            agent2.output)
        self.runs.append(agent3)

        # Agent 4 stops with a message when the input holds fewer than two
        # suppliers, which is a legitimate answer on a small extract rather than
        # a failure of the build. Its output is a companion file, so the wide
        # table is complete whether or not it ran.
        agent4 = self._invoke(
            AgentRun("Agent 4 - supplier consolidation", "agent4.py",
                     self.work / "agent4_supplier_consolidation.csv",
                     tag="Agent 4"),
            [*common, "--input", str(agent2.output),
             "--registry", str(self.here / "lexicon" / "agent4_supplier_registry.json")],
            agent2.output)
        self.runs.append(agent4)

    # -- merging -------------------------------------------------------------

    def _index(self, run: AgentRun) -> Optional[Dict[int, Dict[str, str]]]:
        """Read one agent's output, keyed by the stage-3 row it describes.

        Only the columns that agent introduces are kept. The rest of its table
        either repeats what an earlier agent already merged or restates a
        business key the wide row carries under the client's own header, and
        holding all of it for every row would cost memory for nothing.
        """
        with run.output.open("r", encoding=CSV_ENCODING, newline="") as handle:
            reader = csv.DictReader(handle)
            headers = [name for name in (reader.fieldnames or []) if name]
            missing = [name for name in run.introduced if name not in headers]
            if AGENT_ROW_KEY not in headers:
                run.ok = False
                run.reason = (f"{run.output.name} has no {AGENT_ROW_KEY} column, so its "
                              f"rows cannot be traced back to a transaction")
                LOGGER.error("%s: %s", run.name, run.reason)
                return None
            if missing:
                # Stopping is the right answer. These names are the deliverable,
                # and a wide table quietly missing a promised column is worse
                # than a run that failed and said which one.
                run.ok = False
                run.reason = (f"{run.output.name} is missing promised column(s): "
                              f"{', '.join(missing)}")
                LOGGER.error("%s: %s", run.name, run.reason)
                return None

            wanted = tuple(run.introduced)
            indexed: Dict[int, Dict[str, str]] = {}
            duplicates = 0
            for record in reader:
                raw = (record.get(AGENT_ROW_KEY) or "").strip()
                try:
                    position = int(raw) - AGENT_ROW_OFFSET
                except ValueError:
                    continue
                if position in indexed:
                    duplicates += 1
                    continue
                indexed[position] = {name: (record.get(name) or "") for name in wanted}
            run.rows = len(indexed)

        if duplicates:
            run.ok = False
            run.reason = (f"{duplicates} row(s) of {run.output.name} point at a stage-3 "
                          f"row already claimed by another")
            LOGGER.error("%s: %s", run.name, run.reason)
            return None
        return indexed

    def merge(self, base_csv: Path, base_columns: Sequence[str], run: AgentRun,
              stage_name: str, extra: Optional[Dict[str, Dict[str, str]]] = None,
              extra_columns: Sequence[str] = ()) -> Optional[Tuple[Path, List[str]]]:
        """Widen one table with one agent's columns, writing CSV and JSONL.

        Where a column name is produced by both sides the agent's value is taken,
        but only when it is populated: an agent that could not resolve a field
        must not blank a value Max had already established.
        """
        indexed = self._index(run)
        if indexed is None:
            return None

        added = [name for name in run.introduced if name not in base_columns]
        added += [name for name in extra_columns if name not in base_columns
                  and name not in added]
        overwritten = [name for name in run.introduced if name in base_columns]
        columns = list(base_columns) + added

        results = self.settings.results_dir
        csv_path = results / f"{stage_name}.csv"
        jsonl_path = results / f"{stage_name}.jsonl"

        rows_in = rows_out = matched = 0
        handles: List[Any] = []
        try:
            handle = csv_path.open("w", encoding=CSV_ENCODING, newline="")
            handles.append(handle)
            writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
            writer.writeheader()

            lines = None
            if self.settings.write_jsonl:
                lines = jsonl_path.open("w", encoding="utf-8")
                handles.append(lines)

            for position, row in enumerate(read_csv_rows(base_csv)):
                rows_in += 1
                block = indexed.get(position)
                output = dict(row)
                if block is not None:
                    matched += 1
                    for name, value in block.items():
                        # An empty answer from the agent leaves whatever was
                        # already there; see AGENT_PRECEDENCE_NOTE.
                        if value != "" or name not in output:
                            output[name] = value
                else:
                    for name in added:
                        output.setdefault(name, "")
                if extra is not None:
                    output.update(extra.get(position, {}))
                for name in extra_columns:
                    output.setdefault(name, "")

                writer.writerow(output)
                if lines is not None:
                    lines.write(json.dumps(output, ensure_ascii=False) + "\n")
                rows_out += 1
        finally:
            for opened in handles:
                opened.close()

        if rows_in != rows_out:
            LOGGER.error("%s changed the row count: %d in, %d out.",
                         stage_name, rows_in, rows_out)
            run.ok = False
            run.reason = f"row count changed during the merge ({rows_in} -> {rows_out})"
            return None

        self.statistics[f"{stage_name}_rows"] = rows_out
        self.statistics[f"{stage_name}_annotated"] = matched
        if overwritten:
            self.statistics[f"{stage_name}_overridden_columns"] = sorted(overwritten)
        self.outputs.append(csv_path.name)
        if self.settings.write_jsonl:
            self.outputs.append(jsonl_path.name)

        LOGGER.info("%s: %d row(s), %d annotated by %s, %d column(s) added.",
                    stage_name, rows_out, matched, run.name, len(added))
        if matched != rows_out:
            LOGGER.warning("%s: %d row(s) received no values from %s.",
                           stage_name, rows_out - matched, run.name)
        return csv_path, columns

    # -- Agent 4 -------------------------------------------------------------

    def _read_supplier_keys(self) -> Dict[str, str]:
        """Map every supplier spelling and id Agent 4 saw onto its resolved key."""
        master = self.work / "agent4_supplier_master.csv"
        if not master.is_file():
            return {}
        keys: Dict[str, str] = {}
        with master.open("r", encoding=CSV_ENCODING, newline="") as handle:
            for record in csv.DictReader(handle):
                key = (record.get("Supplier_Key") or "").strip()
                if not key:
                    continue
                spellings = [record.get("Canonical_Supplier_Name") or ""]
                spellings += (record.get("Raw_Name_Variants") or "").split(";")
                spellings += (record.get("Supplier_Ids") or "").split(";")
                for spelling in spellings:
                    token = lookup_key(spelling)
                    if token:
                        keys.setdefault(token, key)
        return keys

    def supplier_key_block(self, base_csv: Path) -> Dict[int, Dict[str, str]]:
        """The resolved supplier key for each row, ready to merge.

        Matched on the supplier name or id already on the row, using the same
        resolution Agent 4 published, so the wide table and the consolidation
        file cannot disagree about which rows belong to which company.
        """
        self.supplier_keys = self._read_supplier_keys()
        if not self.supplier_keys:
            return {}

        block: Dict[int, Dict[str, str]] = {}
        resolved = 0
        for position, row in enumerate(read_csv_rows(base_csv)):
            key = ""
            for column in ("Supplier_Name", "Supplier", "ERP supplier name",
                           "Supplier_Id", "ERP supplier number"):
                token = lookup_key(row.get(column, ""))
                if token and token in self.supplier_keys:
                    key = self.supplier_keys[token]
                    break
            block[position] = {SUPPLIER_KEY_COLUMN: key}
            if key:
                resolved += 1
        self.statistics["supplier_keys_resolved"] = resolved
        return block

    def copy_consolidation(self) -> None:
        """Publish Agent 4's analysis beside the wide table.

        Copied rather than merged: one row here is a supplier within a comparison
        scope, and folding that onto a purchase line would either multiply the
        rows or invent a summary nobody asked for. Joined to the wide table on
        the supplier key plus the scope.
        """
        results = self.settings.results_dir
        for source_name, target_name in (
                ("agent4_supplier_consolidation.csv", "max_supplier_consolidation.csv"),
                ("agent4_supplier_consolidation.jsonl", "max_supplier_consolidation.jsonl"),
                ("agent4_supplier_pairs.csv", "max_supplier_pairs.csv"),
                ("agent4_supplier_master.csv", "max_supplier_master.csv")):
            source = self.work / source_name
            if not source.is_file():
                continue
            shutil.copyfile(source, results / target_name)
            self.outputs.append(target_name)

    # -- entry point ---------------------------------------------------------

    def build(self, stage3_csv: Path,
              stage3_columns: Sequence[str]) -> Tuple[Optional[Path], List[str]]:
        """Run the agents and write stages 4, 5 and 6. Returns the last stage."""
        self.run_agents(stage3_csv)
        by_script = {run.script: run for run in self.runs}

        current_csv: Optional[Path] = None
        current_columns = list(stage3_columns)
        base_csv = stage3_csv

        plan = (("agent1.py", "max_stage4_enriched"),
                ("agent2.py", "max_stage5_grouped"),
                ("agent3.py", "max_stage6_standardised"))

        for script, stage_name in plan:
            run = by_script.get(script)
            if run is None or not run.ok:
                break
            # The supplier key rides along with the last merge, so that the
            # column exists on the table a reader will actually open.
            extra: Optional[Dict[str, Dict[str, str]]] = None
            extra_columns: Sequence[str] = ()
            if script == "agent3.py" and by_script.get("agent4.py", AgentRun("", "", Path())).ok:
                extra = self.supplier_key_block(base_csv)
                extra_columns = (SUPPLIER_KEY_COLUMN,)

            merged = self.merge(base_csv, current_columns, run, stage_name,
                                extra=extra, extra_columns=extra_columns)
            if merged is None:
                break
            base_csv, current_columns = merged
            current_csv = base_csv

        if by_script.get("agent4.py", AgentRun("", "", Path())).ok:
            self.copy_consolidation()

        self.statistics["agents"] = {
            run.name: ("ok" if run.ok else (run.reason or "not run"))
            for run in self.runs
        }
        return current_csv, current_columns

    def missing_columns(self, columns: Sequence[str]) -> List[str]:
        """Promised columns that did not reach the final table."""
        promised = list(AGENT1_REQUIRED) + list(AGENT2_REQUIRED) + list(AGENT3_REQUIRED)
        present = set(columns)
        return [name for name in promised if name not in present]

    def explain_missing(self, missing: Sequence[str]) -> None:
        """Say why columns are absent and what to do, not just which ones.

        A bare list of fifty-six names describes the damage without naming the
        cause. The agents run in a chain, so one agent stopping accounts for its
        own columns and for every column the agents after it would have added, and
        the first agent to stop is the only thing worth acting on.
        """
        blocked = next((run for run in self.runs if not run.ok and run.reason), None)
        LOGGER.error("%d promised column(s) are missing from the final table.",
                     len(missing))

        if blocked is None:
            LOGGER.error("The agents were not run, so none of their columns exist. "
                         "Run without --no-agents to add them.")
            return

        after = [run.name for run in self.runs
                 if run.ok is False and run is not blocked]
        LOGGER.error("Cause: %s %s.", blocked.name, blocked.reason)
        if after:
            LOGGER.error("The agents after it did not run, so their columns are "
                         "missing too: %s.", ", ".join(after))
        if blocked.log and blocked.log.is_file():
            LOGGER.error("Its full log: %s", blocked.log)

        finished = [run for run in self.runs if run.ok]
        if finished:
            LOGGER.error("Kept, and reused rather than recomputed on the next run: "
                         "%s.", ", ".join(run.name for run in finished))
        LOGGER.error("Rerun the same command to carry on from %s.", blocked.name)


class ResultsLock:
    """Stops two builds writing into one results folder at the same time.

    Every stage writes to a fixed name, so two builds sharing a folder do not
    collide loudly: they interleave, each overwrites the other's intermediates,
    and the dataset that survives is whichever build finished last. It looks
    complete and is a mixture of two configurations. Kept in step with the class
    of the same name in all_agents.py, which shares the folder.
    """

    NAME = ".build.lock"

    def __init__(self, results_dir: Path) -> None:
        self.path = results_dir / self.NAME
        self.held = False

    def _holder(self) -> Dict[str, Any]:
        try:
            return json.loads(self.path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            return {}

    @staticmethod
    def _alive(pid: int) -> bool:
        """Whether a process is still running, without disturbing it."""
        if pid <= 0:
            return False
        try:
            os.kill(pid, 0)  # asks the kernel, delivers nothing
        except ProcessLookupError:
            return False
        except PermissionError:
            return True  # somebody else's, but running
        except OSError:
            return False
        return True

    def acquire(self) -> Optional[str]:
        """Take the lock, or return why it cannot be taken."""
        self.path.parent.mkdir(parents=True, exist_ok=True)

        if self.path.exists():
            holder = self._holder()
            pid = int(holder.get("pid") or 0)
            if self._alive(pid) and pid != os.getpid():
                started = holder.get("started") or "an unknown time"
                return (f"another build is already writing to this folder: "
                        f"process {pid}, started {started}.\n"
                        f"  Wait for it to finish, or stop it, or point this build "
                        f"at a different --results folder.\n"
                        f"  If that process is gone, delete {self.path}.")
            if pid and pid != os.getpid():
                LOGGER.warning("Taking over the lock left by process %d, which is "
                               "no longer running.", pid)

        self.path.write_text(json.dumps({
            "pid": os.getpid(),
            "started": datetime.datetime.now().isoformat(timespec="seconds"),
            "command": " ".join(sys.argv),
        }, indent=2), encoding="utf-8")
        self.held = True
        return None

    def release(self) -> None:
        """Give the lock up, but never somebody else's."""
        if not self.held:
            return
        self.held = False
        if int(self._holder().get("pid") or 0) == os.getpid():
            self.path.unlink(missing_ok=True)


class RunJournal:
    """What a run has finished, so a later run can carry on from there.

    Stage 3 reads every distinct string in the extract and the agents after it run
    for hours, so a build stopped part way through has usually earned most of what
    it was going to. Starting again from stage 1 throws that away, and on a full
    extract that is most of a day.

    A stage is recorded only once its files are closed, together with their size
    and modification time. That is what separates a finished stage from an
    interrupted one: a half-written CSV is never in here, and one that has been
    touched since no longer matches and is rebuilt. The recorded statistics are
    kept too, because the summary and the row-count check at the end of the run
    need what a skipped stage would have measured.
    """

    NAME = ".max_run_journal.json"
    VERSION = "1"

    def __init__(self, settings: "Settings") -> None:
        self.settings = settings
        self.path = settings.results_dir / self.NAME
        self.payload: Dict[str, Any] = {}

    # -- deciding whether the record still applies ---------------------------

    def fingerprint(self) -> Dict[str, Any]:
        """What must be unchanged for finished stages to still be valid.

        The extracts are compared by size and modification time rather than by
        content: reading every byte of the source folder to decide whether reading
        it can be skipped would defeat the point, and a changed file is what needs
        detecting, not a changed byte.
        """
        extracts = []
        if self.settings.source_dir.is_dir():
            for path in sorted(self.settings.source_dir.rglob("*")):
                if (path.is_file() and not path.name.startswith((".", "~$"))
                        and path.suffix.lower() in {".csv", ".xlsx", ".xls"}):
                    try:
                        stat = path.stat()
                    except OSError:
                        continue
                    extracts.append({"file": str(path.relative_to(self.settings.source_dir)),
                                     "bytes": stat.st_size,
                                     "modified": int(stat.st_mtime)})
        return {
            "extracts": extracts,
            "lexicon": _digest(self.settings.lexicon_path),
            # Only the settings that change what stages 1 to 3 write. The agent
            # settings are absent on purpose: they cannot alter these stages, and
            # including them would rebuild the table whenever an agent flag moved.
            "settings": {
                "native_po_columns": self.settings.native_po_columns,
                "interpretation_floor": self.settings.interpretation_floor,
                "min_text_length": self.settings.min_text_length,
                "write_jsonl": self.settings.write_jsonl,
                "model": (self.settings.model.model
                          if self.settings.model.enabled else ""),
            },
        }

    def load(self) -> Optional[str]:
        """Read the record, returning why it cannot be used, or None if it can."""
        if not self.path.is_file():
            return "no previous run"
        try:
            payload = json.loads(self.path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            return "the record of the previous run is unreadable"
        if payload.get("version") != self.VERSION:
            return "the record was written by a different version of Max"
        if payload.get("state") == "complete":
            return "the previous run finished"
        if payload.get("fingerprint") != self.fingerprint():
            return self._what_changed(payload.get("fingerprint") or {})
        self.payload = payload
        return None

    def _what_changed(self, before: Dict[str, Any]) -> str:
        """Name the change that invalidates the record, rather than just its fact."""
        now = self.fingerprint()
        if before.get("settings") != now["settings"]:
            moved = [name for name, value in now["settings"].items()
                     if (before.get("settings") or {}).get(name) != value]
            return f"the run's settings changed ({', '.join(moved) or 'unknown'})"
        if before.get("lexicon") != now["lexicon"]:
            return "the controlled vocabulary changed"

        was = {entry["file"]: entry for entry in before.get("extracts") or []}
        is_ = {entry["file"]: entry for entry in now["extracts"]}
        added = sorted(set(is_) - set(was))
        removed = sorted(set(was) - set(is_))
        changed = sorted(name for name in set(was) & set(is_) if was[name] != is_[name])
        parts = []
        if added:
            parts.append(f"{len(added)} extract(s) added ({', '.join(added[:2])})")
        if removed:
            parts.append(f"{len(removed)} removed ({', '.join(removed[:2])})")
        if changed:
            parts.append(f"{len(changed)} changed ({', '.join(changed[:2])})")
        return "the source extracts changed: " + ("; ".join(parts) or "unknown")

    # -- what the previous run got through -----------------------------------

    def finished(self, stage: str) -> Optional[Dict[str, Any]]:
        """A stage's record, if it completed and its files are still as written."""
        record = ((self.payload.get("stages") or {}).get(stage) or {})
        if not record.get("done"):
            return None
        for described in record.get("files") or []:
            path = self.settings.results_dir / described["file"]
            try:
                stat = path.stat()
            except OSError:
                LOGGER.info("%s cannot be reused: %s is gone.", stage, described["file"])
                return None
            if (stat.st_size != described.get("bytes")
                    or int(stat.st_mtime) != described.get("modified")):
                LOGGER.info("%s cannot be reused: %s has changed since it was written.",
                            stage, described["file"])
                return None
        return record

    def interrupted_at(self) -> str:
        return str(self.payload.get("running") or "")

    def done_stages(self) -> List[str]:
        """Finished stages in the order they run, not the order they were stored."""
        done = {name for name, record in (self.payload.get("stages") or {}).items()
                if record.get("done")}
        return [name for name in STAGE_ORDER if name in done]

    # -- recording -----------------------------------------------------------

    def begin(self, run_id: str) -> None:
        stages = self.payload.get("stages") or {}
        self.payload = {
            "version": self.VERSION,
            "run_id": run_id,
            "started": self.payload.get("started") or _now(),
            "fingerprint": self.fingerprint(),
            "state": "running",
            "running": "",
            "stages": stages,
        }
        self._save()

    def starting(self, stage: str) -> None:
        self.payload["running"] = stage
        self._save()

    def completed(self, stage: str, outputs: Sequence[str],
                  statistics: Dict[str, Any], diagnostics: Dict[str, Any],
                  extra: Optional[Dict[str, Any]] = None) -> None:
        """Record a stage whose files are written and closed."""
        described = []
        for name in outputs:
            path = self.settings.results_dir / name
            try:
                stat = path.stat()
            except OSError:
                continue
            described.append({"file": name, "bytes": stat.st_size,
                              "modified": int(stat.st_mtime)})
        self.payload.setdefault("stages", {})[stage] = {
            "done": True,
            "at": _now(),
            "files": described,
            "statistics": copy.deepcopy(statistics),
            "diagnostics": copy.deepcopy(diagnostics),
            **(extra or {}),
        }
        self.payload["running"] = ""
        self._save()

    def stopped(self, stage: str) -> None:
        self.payload["state"] = "interrupted"
        self.payload["running"] = stage
        self.payload["stopped_at"] = _now()
        self._save()

    def complete(self) -> None:
        self.payload["state"] = "complete"
        self.payload["running"] = ""
        self.payload["finished_at"] = _now()
        self._save()

    def discard(self) -> None:
        self.payload = {}
        try:
            self.path.unlink()
        except OSError:
            pass

    def _save(self) -> None:
        try:
            self.settings.results_dir.mkdir(parents=True, exist_ok=True)
            self.payload["updated"] = _now()
            self.path.write_text(
                json.dumps(self.payload, ensure_ascii=False, indent=2, sort_keys=True),
                encoding="utf-8")
        except OSError as error:
            LOGGER.debug("Could not record run progress: %s", error)


class Builder:
    """Runs the three stages and writes their outputs."""

    def __init__(self, settings: Settings) -> None:
        self.settings = settings
        self.lexicon = Lexicon.load(settings.lexicon_path)
        self.interpreter = TextInterpreter(self.lexicon, settings)

        self.model: Optional[LanguageModelClient] = None
        if settings.model.enabled:
            self.model = LanguageModelClient(
                settings.model, settings.cache_dir / "max_model_cache.json",
                interactive=settings.interactive)

        self.run_id = ""
        self.sievo_tables: List[Table] = []
        self.invoice_index: Optional[InvoiceIndex] = None
        self.po_joiner: Optional[PurchaseOrderJoiner] = None
        self.sievo_headers: List[str] = []
        self.sievo_map: Optional[ColumnMap] = None
        self.outputs: List[str] = []
        self.statistics: Dict[str, Any] = {}
        self.diagnostics: Dict[str, Any] = {}

        # Stage 3's header, kept so that the agent stages know what the table
        # already carries and can tell an added column from an overridden one.
        self.stage3_columns: List[str] = []
        self.pipeline: Optional[AgentPipeline] = None
        self.final_csv: Optional[Path] = None
        self.final_columns: List[str] = []

        # Purchase-order extracts that were read but cannot be joined. Held
        # separately because the joiner discards them, and the diagnostics file
        # is precisely where someone needs to see that a delivered file was
        # unusable.
        self.unusable_po_systems: List[str] = []

        # Stages taken from an interrupted run rather than run again. Reported, so
        # that a table nobody watched being built still says where it came from.
        self.carried_over: List[str] = []

    # -- setup --------------------------------------------------------------

    def load(self) -> None:
        grouped = discover_tables(self.settings.source_dir)

        self.sievo_tables = grouped.get("sievo") or []
        if not self.sievo_tables:
            raise SystemExit(
                f"No transaction extract was found under {self.settings.source_dir}.\n"
                "Max builds outwards from the Sievo transaction table; without it there "
                "is nothing to widen. Expected a file carrying columns such as "
                "SourceRowId, DataSource, PO number and Spend in EUR.")

        # One header list covering every transaction table supplied, in the
        # order the columns were first seen.
        headers: List[str] = []
        for table in self.sievo_tables:
            for header in table.headers:
                if header and header not in headers:
                    headers.append(header)
        self.sievo_headers = headers
        self.sievo_map = ColumnMap(headers, SIEVO_FIELDS)

        invoice_tables = grouped.get("invoice") or []
        if invoice_tables:
            self.invoice_index = InvoiceIndex(invoice_tables)
            LOGGER.info("Invoice extract: %d rows (%d line rows, %d summary rows, "
                        "%d duplicates dropped).",
                        len(self.invoice_index.lines),
                        sum(1 for line in self.invoice_index.lines if line.is_line),
                        self.invoice_index.summary_rows, self.invoice_index.duplicates)
        else:
            LOGGER.warning("No invoice extract found; stage 1 columns will be empty.")

        indexes: Dict[str, PurchaseOrderIndex] = {}
        for name, synonyms in (("maximo", MAXIMO_FIELDS), ("basware", BASWARE_FIELDS)):
            tables = grouped.get(name) or []
            if not tables:
                continue
            index = PurchaseOrderIndex(name, tables, synonyms)
            if index.empty:
                self.unusable_po_systems.append(name)
                LOGGER.warning("%s extract carries no usable PO number; "
                               "it cannot take part in the join.", name.capitalize())
            else:
                LOGGER.info("%s extract: %d PO lines across %d orders.",
                            name.capitalize(), len(index.records), len(index._by_order))
            indexes[name] = index

        if not any(not index.empty for index in indexes.values()):
            LOGGER.warning("No purchase-order extract carries a usable PO number; "
                           "stage 2 columns will be empty.")
        self.po_joiner = PurchaseOrderJoiner(indexes, self.settings.native_po_columns)

        self.run_id = stable_hash(
            AGENT_VERSION, self.lexicon.version,
            *(table.label for table in self.sievo_tables))

    # -- stages 1 and 2 -----------------------------------------------------

    def build_wide(self) -> Tuple[Path, Path, List[str]]:
        """Stream the transactions, widening each row, writing stages 1 and 2.

        Both stages are produced in a single pass. Stage 1's file is the
        transaction plus the invoice block; stage 2's is that same row plus the
        purchase-order block, so writing them together costs one pass rather
        than two and cannot let the two files disagree.
        """
        results = self.settings.results_dir
        results.mkdir(parents=True, exist_ok=True)

        stage1_columns = ["Max_Row_Id"] + self.sievo_headers + INVOICE_COLUMNS
        stage2_columns = stage1_columns + self.po_joiner.columns

        stage1_csv = results / "max_stage1_sievo_invoice.csv"
        stage2_csv = results / "max_stage2_with_po.csv"
        stage1_jsonl = results / "max_stage1_sievo_invoice.jsonl"
        stage2_jsonl = results / "max_stage2_with_po.jsonl"

        counters: Counter = Counter()
        invoice_strategies: Counter = Counter()
        po_strategies: Counter = Counter()
        rows_in = 0

        handles = []
        try:
            stage1_handle = stage1_csv.open("w", encoding=CSV_ENCODING, newline="")
            stage2_handle = stage2_csv.open("w", encoding=CSV_ENCODING, newline="")
            handles += [stage1_handle, stage2_handle]
            stage1_writer = csv.DictWriter(stage1_handle, fieldnames=stage1_columns,
                                           extrasaction="ignore")
            stage2_writer = csv.DictWriter(stage2_handle, fieldnames=stage2_columns,
                                           extrasaction="ignore")
            stage1_writer.writeheader()
            stage2_writer.writeheader()

            stage1_lines = stage2_lines = None
            if self.settings.write_jsonl:
                stage1_lines = stage1_jsonl.open("w", encoding="utf-8")
                stage2_lines = stage2_jsonl.open("w", encoding="utf-8")
                handles += [stage1_lines, stage2_lines]

            for table in self.sievo_tables:
                for row_number, record in table.iter_records():
                    rows_in += 1
                    row = {header: normalise_text(record.get(header, ""))
                           for header in self.sievo_headers}
                    keys = {name: self.sievo_map.get(record, name) for name in SIEVO_FIELDS}

                    row_id = keys.get("row_id") or ""
                    row["Max_Row_Id"] = stable_hash(table.label, str(row_number), row_id)

                    invoice_block = join_invoice(keys, self.invoice_index)
                    counters[f"invoice_{invoice_block['Invoice_Match_Level']}"] += 1
                    if invoice_block["Invoice_Match_Strategy"]:
                        invoice_strategies[invoice_block["Invoice_Match_Strategy"]] += 1

                    stage1_row = {**row, **invoice_block}
                    stage1_writer.writerow(stage1_row)
                    if stage1_lines is not None:
                        stage1_lines.write(json.dumps(stage1_row, ensure_ascii=False) + "\n")

                    po_block = self.po_joiner.join(keys)
                    counters[f"po_{po_block['PO_Match_Level']}"] += 1
                    if po_block["PO_Match_Strategy"]:
                        po_strategies[po_block["PO_Match_Strategy"]] += 1

                    stage2_row = {**stage1_row, **po_block}
                    stage2_writer.writerow(stage2_row)
                    if stage2_lines is not None:
                        stage2_lines.write(json.dumps(stage2_row, ensure_ascii=False) + "\n")
        finally:
            for handle in handles:
                handle.close()

        self.statistics.update({
            "rows_in": rows_in,
            "invoice_matched_line": counters["invoice_line"],
            "invoice_matched_header": counters["invoice_header"],
            "invoice_unmatched": counters["invoice_none"],
            "po_matched_line": counters["po_line"],
            "po_matched_header": counters["po_header"],
            "po_unmatched": counters["po_none"],
        })
        self.diagnostics["invoice_strategies"] = dict(invoice_strategies)
        self.diagnostics["po_strategies"] = dict(po_strategies)

        self.outputs += [stage1_csv.name, stage2_csv.name]
        if self.settings.write_jsonl:
            self.outputs += [stage1_jsonl.name, stage2_jsonl.name]
        return stage2_csv, stage2_jsonl, stage2_columns

    # -- stage 3 ------------------------------------------------------------

    def interpret(self, stage2_csv: Path, stage2_columns: List[str]) -> Path:
        """Read the free text on every row of the wide table.

        Two passes over the file on disk rather than one over the rows in
        memory: the first collects the distinct strings so the model tier can be
        offered a de-duplicated worklist, the second writes the output. At a
        million rows the distinct strings number in the tens of thousands, which
        is the difference between a run that is affordable and one that is not.
        """
        results = self.settings.results_dir
        stage3_csv = results / "max_stage3_interpreted.csv"
        stage3_jsonl = results / "max_stage3_interpreted.jsonl"
        columns = stage2_columns + Interpretation.columns()
        self.stage3_columns = list(columns)

        distinct: Counter = Counter()
        for row in self._read_rows(stage2_csv):
            describing, supporting, _ = self._gather_text(row)
            if describing or supporting:
                distinct[(describing, supporting)] += 1
        LOGGER.info("Stage 3: %d distinct free-text combinations to read.", len(distinct))

        # Interpret every distinct combination once, then decide which readings
        # are weak enough to be worth a model call.
        readings: Dict[Tuple[str, str], Interpretation] = {
            key: self.interpreter.interpret(key[0], key[1], "")
            for key in sorted(distinct)
        }

        model_answers: Dict[str, Dict[str, str]] = {}
        if self.model is not None:
            # The model is asked about the descriptive text only, and asked once
            # per distinct string: two rows whose notes differ but whose
            # description is the same are one question, not two.
            residue: List[str] = []
            contexts: Dict[str, str] = {}
            for key, reading in readings.items():
                describing, supporting = key
                subject = describing or supporting
                if not subject:
                    continue
                contexts[subject] = supporting if describing else ""
                if self._needs_model(reading, describing, supporting):
                    residue.append(subject)
            residue = sorted(set(residue))
            LOGGER.info("Stage 3: %d of %d readings need the model "
                        "(%d below the confidence floor, remainder leftover "
                        "non-English or mixed evidence).",
                        len(residue), len(readings),
                        sum(1 for reading in readings.values()
                            if reading.confidence < self.settings.interpretation_floor))
            reader = ModelReader(self.model, self.settings)
            if residue:
                model_answers = reader.resolve(residue, contexts)

            # Second pass: anything that still is not English, or whose extra
            # evidence looks like a different purchase, is reviewed explicitly.
            review_texts: List[str] = []
            for key, reading in readings.items():
                describing, supporting = key
                subject = describing or supporting
                if not subject:
                    continue
                preview = ((model_answers.get(subject) or {}).get("description")
                           or reading.description)
                if (foreign_tokens_in(preview)
                        or self._looks_unrelated(describing, supporting)):
                    review_texts.append(subject)
            review_texts = sorted(set(review_texts))
            if review_texts:
                LOGGER.info("Stage 3: reviewing %d description(s) for English "
                            "and relevance.", len(review_texts))
                model_answers.update(
                    reader.resolve(review_texts, contexts, review=True))

        weak = sum(1 for reading in readings.values()
                   if reading.confidence < self.settings.interpretation_floor)
        rows_out = 0
        handles = []
        try:
            handle = stage3_csv.open("w", encoding=CSV_ENCODING, newline="")
            handles.append(handle)
            writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
            writer.writeheader()

            lines = None
            if self.settings.write_jsonl:
                lines = stage3_jsonl.open("w", encoding="utf-8")
                handles.append(lines)

            for row in self._read_rows(stage2_csv):
                describing, supporting, sources = self._gather_text(row)
                reading = readings.get((describing, supporting))
                if reading is None:
                    block = Interpretation(sources=sources).as_columns()
                else:
                    block = self._apply(reading, sources,
                                        model_answers.get(describing or supporting))
                output = {**row, **block}
                writer.writerow(output)
                if lines is not None:
                    lines.write(json.dumps(output, ensure_ascii=False) + "\n")
                rows_out += 1
        finally:
            for opened in handles:
                opened.close()

        self.statistics.update({
            "distinct_texts": len(distinct),
            "texts_below_floor": weak,
            "texts_read_by_model": len(model_answers),
            "rows_out": rows_out,
        })
        self.outputs.append(stage3_csv.name)
        if self.settings.write_jsonl:
            self.outputs.append(stage3_jsonl.name)
        self.final_csv = stage3_csv
        self.final_columns = list(columns)
        return stage3_csv

    # -- stages 4 to 6 ------------------------------------------------------

    # -- stages 1 to 3, run or carried over ----------------------------------

    def wide_stage(self, journal: RunJournal) -> Tuple[Path, Path, List[str]]:
        """Stages 1 and 2, or what a previous run already wrote for them."""
        results = self.settings.results_dir
        stage1_csv = results / "max_stage1_sievo_invoice.csv"
        stage2_csv = results / "max_stage2_with_po.csv"

        record = journal.finished("wide") if self.settings.resume else None
        if record:
            columns = _header_of(stage2_csv)
            if columns:
                self._carry_over(record, "Stages 1 and 2")
                self.carried_over.append("stages 1 and 2 (invoice and PO joins)")
                return stage2_csv, stage1_csv, columns
            LOGGER.info("Stages 1 and 2 cannot be reused: %s has no header.",
                        stage2_csv.name)

        journal.starting("wide")
        stage2_csv, stage1_csv, columns = self.build_wide()
        journal.completed("wide", self.outputs, self.statistics, self.diagnostics)
        return stage2_csv, stage1_csv, columns

    def interpret_stage(self, journal: RunJournal, stage2_csv: Path,
                        stage2_columns: List[str]) -> Path:
        """Stage 3, or what a previous run already wrote for it."""
        stage3_csv = self.settings.results_dir / "max_stage3_interpreted.csv"

        record = journal.finished("interpret") if self.settings.resume else None
        if record:
            columns = _header_of(stage3_csv)
            if columns:
                self._carry_over(record, "Stage 3")
                self.stage3_columns = columns
                self.final_csv = stage3_csv
                self.final_columns = list(columns)
                self.carried_over.append("stage 3 (free text read into columns)")
                return stage3_csv
            LOGGER.info("Stage 3 cannot be reused: %s has no header.", stage3_csv.name)

        journal.starting("interpret")
        stage3_csv = self.interpret(stage2_csv, stage2_columns)
        journal.completed("interpret", self.outputs, self.statistics, self.diagnostics)
        return stage3_csv

    def _carry_over(self, record: Dict[str, Any], label: str) -> None:
        """Adopt a finished stage's measurements as if it had just run.

        The summary and the row-count check at the end of the run read these, so a
        stage that is skipped has to hand back what it measured or the run reports
        a table of no rows and then calls that a failure.
        """
        self.statistics.update(record.get("statistics") or {})
        self.diagnostics.update(record.get("diagnostics") or {})
        for described in record.get("files") or []:
            if described["file"] not in self.outputs:
                self.outputs.append(described["file"])
        LOGGER.info("%s already done on %s; carrying it over.",
                    label, record.get("at") or "an earlier run")

    def enrich(self, stage3_csv: Path) -> Optional[Path]:
        """Run the agents and widen the table with what they found.

        Nothing here can invalidate stages 1 to 3: those files are already on
        disk and are what they were. An agent that fails costs the columns it
        would have added and is reported as having failed, and the stage before it
        remains the table to use.
        """
        self.pipeline = AgentPipeline(self.settings)
        final_csv, final_columns = self.pipeline.build(stage3_csv, self.stage3_columns)

        self.outputs += self.pipeline.outputs
        self.statistics["agent_stages"] = self.pipeline.statistics

        if final_csv is not None:
            self.final_csv = final_csv
            self.final_columns = final_columns

        missing = self.pipeline.missing_columns(self.final_columns)
        if missing:
            self.statistics["columns_missing"] = missing
            self.pipeline.explain_missing(missing)
        return final_csv

    def _apply(self, reading: Interpretation, sources: str,
               answer: Optional[Dict[str, str]]) -> Dict[str, str]:
        """Combine a local reading with the model's answer, if there is one.

        The model replaces the prose fields. Extracted values stay as the rules
        found them, because those are exact and the model's are not. Foreign
        common nouns are stripped from every published prose column so a Finnish
        leftover cannot reach the output even when the model is off.
        """
        result = Interpretation(**{**reading.__dict__, "sources": sources})
        if answer:
            result.description = answer.get("description") or result.description
            if answer.get("type"):
                result.item_or_service = answer.get("type", "")
            if answer.get("intent"):
                result.intent = answer.get("intent", "")
            result.method = "model"
            result.confidence = max(result.confidence, 0.60)
        result.description = drop_foreign_common_nouns(result.description)
        result.intent = drop_foreign_common_nouns(result.intent)
        result.keywords = self.interpreter._keywords(result.description)
        return result.as_columns()

    def _needs_model(self, reading: Interpretation, describing: str, supporting: str) -> bool:
        """Whether the local reading is not yet a safe English description."""
        if not (describing or supporting):
            return False
        if (reading.confidence < self.settings.interpretation_floor
                and reading.token_count >= 1):
            return True
        if foreign_tokens_in(reading.description) or foreign_tokens_in(describing):
            return True
        if reading.language and reading.language not in {"en", ""}:
            return True
        if reading.method in {"passthrough", "none"} and reading.token_count >= 1:
            return True
        return False

    @staticmethod
    def _looks_unrelated(describing: str, supporting: str) -> bool:
        """True when extra evidence shares no content words with the line."""
        if not describing or not supporting:
            return False
        primary = {lookup_key(token) for token in tokenise(describing)
                   if len(token) > 3 and not any(ch.isdigit() for ch in token)}
        extra = {lookup_key(token) for token in tokenise(supporting)
                 if len(token) > 3 and not any(ch.isdigit() for ch in token)}
        return bool(primary) and bool(extra) and not (primary & extra)

    @staticmethod
    def _read_rows(path: Path) -> Iterator[Dict[str, str]]:
        return read_csv_rows(path)

    @staticmethod
    def _gather_text(row: Dict[str, str]) -> Tuple[str, str, str]:
        """Collect the row's free text, in priority order, without repetition.

        Returns the descriptive text, the supporting text and the list of
        fields that contributed. A filled purchase line is never concatenated
        with a joined invoice article: that mix is how a false join used to
        leak Finnish into an English description.
        """
        primary: List[str] = []
        fallback: List[str] = []
        supporting: List[str] = []
        used: List[str] = []
        seen: Set[str] = set()
        for column, label, role in TEXT_SOURCE_COLUMNS:
            value = normalise_text(row.get(column, ""))
            if not value or is_blank(value):
                continue
            key = lookup_key(value)
            if key in seen:
                continue
            seen.add(key)
            if role == "primary":
                primary.append(value)
            elif role == "fallback":
                fallback.append(value)
            else:
                supporting.append(value)
            used.append(label)
        describing = " | ".join(primary) if primary else " | ".join(fallback)
        extra = list(supporting)
        if primary:
            extra = fallback + extra
        return describing, " | ".join(extra), "; ".join(used)

    # -- reporting ----------------------------------------------------------

    def write_manifest(self) -> Dict[str, Any]:
        """Record what was read, what was produced and how the joins fared."""
        results = self.settings.results_dir

        if self.model is not None:
            self.statistics["token_usage"] = {
                **self.model.usage.as_dict(),
                **self.model.guard.as_dict(),
            }

        diagnostics = {
            "sources": {
                "transactions": [table.label for table in self.sievo_tables],
                "invoice_lines": len(self.invoice_index.lines) if self.invoice_index else 0,
                "invoice_duplicates_dropped": self.invoice_index.duplicates if self.invoice_index else 0,
                "purchase_order_systems": {
                    name: len(index.records)
                    for name, index in sorted((self.po_joiner.indexes if self.po_joiner else {}).items())
                },
                "purchase_order_systems_unusable": sorted(self.unusable_po_systems),
            },
            "key_population": self.diagnostics.get("key_population", {}),
            "invoice_strategies": self.diagnostics.get("invoice_strategies", {}),
            "po_strategies": self.diagnostics.get("po_strategies", {}),
            "advice": self.diagnostics.get("advice", []),
        }

        diagnostics_path = results / "max_join_diagnostics.json"
        diagnostics_path.write_text(
            json.dumps(diagnostics, ensure_ascii=False, indent=2, sort_keys=True),
            encoding="utf-8")
        manifest_path = results / "max_run_manifest.json"
        self.outputs += [diagnostics_path.name, manifest_path.name]

        manifest = {
            "agent": AGENT_NAME,
            "version": AGENT_VERSION,
            "run_id": self.run_id,
            "lexicon_version": self.lexicon.version,
            "source_dir": str(self.settings.source_dir),
            "results_dir": str(results),
            "settings": {
                "native_po_columns": self.settings.native_po_columns,
                "interpretation_floor": self.settings.interpretation_floor,
                "run_agents": self.settings.run_agents,
                "use_llm": self.settings.model.enabled or bool(self.settings.use_llm),
                "model": self.settings.model.model if self.settings.model.enabled else "",
            },
            "statistics": self.statistics,
            "diagnostics": diagnostics,
            "outputs": sorted(set(self.outputs)),
            "final_table": self.final_csv.name if self.final_csv else "",
            "final_columns": len(self.final_columns),
            "column_precedence": AGENT_PRECEDENCE_NOTE,
            "carried_over": list(self.carried_over),
        }
        manifest_path.write_text(
            json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")
        return manifest

    def measure_keys(self) -> None:
        """Record how populated each candidate join key is.

        This is the block to read when a join matched nothing. It states, for
        each key the join would have used, how many transactions carried it and
        how many of those values were also present on the other side.
        """
        # Seeded with every candidate key so that a key which is never populated
        # is reported as zero rather than omitted. "invoice_number: 0" is the
        # single most useful line in this file when a join found nothing, and
        # leaving it out would hide it.
        tracked = ("invoice_number", "document_number", "document_line",
                   "po_number", "po_line", "document_identifier")
        population: Counter = Counter({name: 0 for name in tracked})
        total = 0
        invoice_keys: Set[str] = set()
        if self.invoice_index is not None:
            invoice_keys = set(self.invoice_index._by_document)

        po_keys: Set[str] = set()
        for index in (self.po_joiner.indexes if self.po_joiner else {}).values():
            po_keys |= set(index._by_order)

        matched_invoice: Set[str] = set()
        matched_po: Set[str] = set()

        for table in self.sievo_tables:
            for _, record in table.iter_records():
                total += 1
                keys = {name: self.sievo_map.get(record, name) for name in SIEVO_FIELDS}
                for name in tracked:
                    if not is_blank(keys.get(name)):
                        population[name] += 1

                for name in ("invoice_number", "document_number"):
                    value = keys.get(name, "")
                    if not is_blank(value) and compact_key(value) in invoice_keys:
                        matched_invoice.add(compact_key(value))
                document = document_id_key(keys.get("document_identifier") or "") \
                    or document_id_key(keys.get("invoice_link") or "")
                if document and document in invoice_keys:
                    matched_invoice.add(document)

                order = keys.get("po_number", "")
                if not is_blank(order) and compact_key(order) in po_keys:
                    matched_po.add(compact_key(order))

        self.diagnostics["key_population"] = {
            "transaction_rows": total,
            "populated": {name: count for name, count in sorted(population.items())},
            "invoice_documents_available": len(invoice_keys),
            "invoice_documents_hit": len(matched_invoice),
            "purchase_orders_available": len(po_keys),
            "purchase_orders_hit": len(matched_po),
        }

        advice: List[str] = []
        if self.invoice_index is not None and not matched_invoice:
            advice.append(
                "No transaction key matched any invoice document. Invoice number, "
                "document number and document identifier were all tried. Check that "
                "the invoice extract covers the same period and entities as the "
                "transaction extract, and that 'Invoice number' is populated.")
        if po_keys and not matched_po:
            advice.append(
                "No PO number on the transaction extract appears in the purchase-order "
                "extracts. Check that the PO extracts cover the same companies as the "
                "transactions; the two sets of order numbers do not currently intersect.")
        for name in self.unusable_po_systems:
            advice.append(
                f"The {name} extract was read but carries no PO number on any row, so it "
                f"cannot be joined at all. Re-export it with the key columns populated.")
        if population.get("po_line", 0) < population.get("po_number", 0):
            advice.append(
                "PO line number is less populated than PO number. Transactions without a "
                "line number can only be matched when the order has a single line; the "
                "rest fall back to header-level fields.")
        self.diagnostics["advice"] = advice

    def close(self) -> None:
        if self.model is not None:
            self.model.save_cache()


# ===========================================================================
# Command line
# ===========================================================================

def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="max.py",
        description="Max - build one wide procurement table from the source extracts.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Examples:\n"
            "  python max.py\n"
            "      Prompt for each path in turn and run with the defaults.\n\n"
            "  python max.py --non-interactive --sources ./sources --results ./results\n"
            "      Run unattended with the local reader only.\n\n"
            "  python max.py --non-interactive --use-llm --llm-spend-limit 10\n"
            "      Add the model tier for free text the rules cannot read.\n\n"
            "  python max.py --non-interactive --no-agents\n"
            "      Stop at stage 3, leaving the agents to be run separately.\n\n"
            "  python max.py --non-interactive --results ./results\n"
            "      After a run was stopped, carry on from the stage it reached.\n\n"
            "  python max.py --non-interactive --restart\n"
            "      Ignore what an interrupted run finished and rebuild from stage 1.\n"
        ),
    )

    paths = parser.add_argument_group("paths")
    paths.add_argument("--sources", metavar="DIR", help="folder holding the source extracts")
    paths.add_argument("--results", metavar="DIR", help="folder to write results into")
    paths.add_argument("--lexicon", metavar="FILE", help="controlled vocabulary JSON file")
    paths.add_argument("--cache", metavar="DIR", help="folder for the model response cache")

    output = parser.add_argument_group("output")
    output.add_argument("--no-native-po-columns", action="store_true",
                        help="omit the raw Maximo and Basware columns, keeping the "
                             "harmonised PO block only")
    output.add_argument("--no-jsonl", action="store_true", help="skip the JSONL exports")

    agents = parser.add_argument_group("agents")
    agents.add_argument("--no-agents", action="store_true",
                        help="stop after stage 3 instead of running Agents 1 to 4 "
                             "and widening the table with their columns")
    agents.add_argument("--agent-timeout", metavar="SECONDS", type=int, default=None,
                        help="abandon an agent still running after this long "
                             "(off by default, so a slow agent is left to finish)")
    agents.add_argument("--agent-silence-timeout", metavar="SECONDS", type=int,
                        default=None,
                        help="treat an agent that has said nothing for this long as "
                             f"hung (default {Settings.agent_silence_timeout})")
    agents.add_argument("--force-agents", action="store_true",
                        help="rerun every agent instead of reusing a result already "
                             "built from the same input")

    resuming = parser.add_argument_group("interrupted runs")
    resuming.add_argument("--restart", action="store_true",
                          help="start from stage 1, discarding the stages an "
                               "interrupted run had already finished")

    tiers = parser.add_argument_group("processing tiers")
    tiers.add_argument("--use-llm", action="store_true",
                       help="let the language model read free text the rules cannot")
    tiers.add_argument("--llm-spend-limit", metavar="USD", type=float, default=None,
                       help="pause and ask once estimated model spend reaches this "
                            f"figure (default {DEFAULT_SPEND_LIMIT:.2f}; 0 disables the alert)")
    tiers.add_argument("--interpretation-floor", type=float, default=0.55,
                       help="confidence below which a reading is offered to the model "
                            "(default 0.55)")

    parser.add_argument("--non-interactive", action="store_true",
                        help="never prompt; use the supplied arguments and defaults")
    parser.add_argument("--verbose", action="store_true", help="emit debug-level logging")
    parser.add_argument("--version", action="version", version=f"{AGENT_NAME} {AGENT_VERSION}")
    return parser


def _clean_path_input(value: str) -> str:
    """Tidy a path pasted into the terminal.

    Dragging a folder onto a terminal wraps it in quotes and escapes its spaces,
    and both would otherwise be taken as part of the name.
    """
    value = value.strip()
    if len(value) >= 2 and value[0] == value[-1] and value[0] in "\"'":
        value = value[1:-1]
    return value.replace("\\ ", " ").strip()


def ask(question: str, default: str) -> str:
    try:
        answer = input(f"{question}\n  [{default}]: ").strip()
    except EOFError:
        return default
    return _clean_path_input(answer) or default


def ask_yes_no(question: str, default: bool) -> bool:
    hint = "Y/n" if default else "y/N"
    try:
        answer = input(f"{question} [{hint}]: ").strip().lower()
    except EOFError:
        return default
    return default if not answer else answer[0] == "y"


def ask_amount(question: str, default: float) -> float:
    """Prompt for a sum of money, re-asking until the answer is usable."""
    while True:
        try:
            answer = input(f"{question}\n  [{default:.2f}]: ").strip()
        except EOFError:
            return default
        if not answer:
            return default
        try:
            value = float(answer.lstrip("$").replace(",", "").strip())
        except ValueError:
            print("  Enter an amount in dollars, for example 25 or 25.00.")
            continue
        if value < 0:
            print("  Enter zero or more; zero runs without a spend alert.")
            continue
        return value


def resolve_settings(args: argparse.Namespace, env: Dict[str, str]) -> Settings:
    here = Path(__file__).resolve().parent
    default_sources = Path(args.sources) if args.sources else here / "sources"
    default_results = Path(args.results) if args.results else here / "results"
    default_lexicon = (Path(args.lexicon) if args.lexicon
                       else here / "lexicon" / "procurement_lexicon.json")
    default_cache = Path(args.cache) if args.cache else here / "cache"

    native = not args.no_native_po_columns
    with_agents = not args.no_agents
    use_llm = args.use_llm
    spend_limit = (args.llm_spend_limit if args.llm_spend_limit is not None
                   else _env_float(env.get("LLM_SPEND_LIMIT"), DEFAULT_SPEND_LIMIT))

    if not args.non_interactive:
        print(BANNER)
        print("\nPress Enter to accept the value shown in brackets.\n")
        default_sources = Path(ask("Source data folder", str(default_sources)))
        default_results = Path(ask("Results folder", str(default_results)))
        default_lexicon = Path(ask("Controlled vocabulary file", str(default_lexicon)))
        default_cache = Path(ask("Cache folder", str(default_cache)))

        print()
        native = ask_yes_no(
            "Carry the raw Maximo and Basware columns into the wide table?", native)
        with_agents = ask_yes_no(
            "Run Agents 1 to 4 and add their columns to the wide table?", with_agents)
        use_llm = ask_yes_no(
            "Let the language model read free text the rules cannot?", use_llm)
        if use_llm:
            print()
            print(f"  Charged at ${INPUT_COST_PER_MTOK:,.2f} per million input tokens and "
                  f"${OUTPUT_COST_PER_MTOK:,.2f} per million output tokens.")
            print("  The run pauses at the figure below and asks before spending more.")
            spend_limit = ask_amount(
                "Alert when estimated language-model spend reaches (USD)", spend_limit)
        print()

    settings = Settings(
        source_dir=default_sources.expanduser().resolve(),
        results_dir=default_results.expanduser().resolve(),
        lexicon_path=default_lexicon.expanduser().resolve(),
        cache_dir=default_cache.expanduser().resolve(),
        native_po_columns=native,
        interpretation_floor=args.interpretation_floor,
        run_agents=with_agents,
        agent_timeout=args.agent_timeout,
        agent_silence_timeout=(args.agent_silence_timeout
                               if args.agent_silence_timeout is not None
                               else Settings.agent_silence_timeout),
        force_agents=args.force_agents or args.restart,
        resume=not args.restart,
        use_llm=use_llm,
        write_jsonl=not args.no_jsonl,
        verbose=args.verbose,
        interactive=not args.non_interactive,
    )
    settings.model = resolve_model_config(env, use_llm, spend_limit)

    if not settings.source_dir.exists():
        raise SystemExit(f"Source folder does not exist: {settings.source_dir}")
    return settings


def configure_logging(verbose: bool) -> None:
    configure_process_logging(verbose)


def print_token_usage(statistics: Dict[str, Any], settings: Settings) -> None:
    """Report language-model consumption for the run."""
    usage = statistics.get("token_usage")
    if not usage:
        return
    print()
    print("-" * 79)
    print("Language model usage")
    print("-" * 79)
    print(f"  Model                : {settings.model.model} ({settings.model.backend})")
    print(f"  Requests sent        : {usage['requests']:,}")
    if usage["failed_requests"]:
        print(f"  Failed requests      : {usage['failed_requests']:,}")
    print(f"  Served from cache    : {usage['cache_hits']:,} (no tokens consumed)")
    print(f"  Input tokens         : {usage['input_tokens']:,}")
    if usage["cached_input_tokens"]:
        print(f"    of which cached    : {usage['cached_input_tokens']:,}")
    print(f"  Output tokens        : {usage['output_tokens']:,}")
    if usage["reasoning_tokens"]:
        print(f"    of which reasoning : {usage['reasoning_tokens']:,}")
    print(f"  Total tokens         : {usage['total_tokens']:,}")

    # Priced from the rates in force at configuration time. Cached input is
    # counted at the full rate, so the figure is an upper bound.
    print(f"  Input cost           : ${usage['input_cost_usd']:,.2f} "
          f"at ${usage['input_cost_per_mtok']:,.2f}/M")
    print(f"  Output cost          : ${usage['output_cost_usd']:,.2f} "
          f"at ${usage['output_cost_per_mtok']:,.2f}/M")
    print(f"  Estimated cost       : ${usage['estimated_cost_usd']:,.2f}")

    if usage.get("spend_limit_usd"):
        print(f"  Spend alert          : ${usage['spend_limit_usd']:,.2f}"
              + (f", raised {usage['spend_limit_extensions']} time(s) "
                 f"to ${usage['spend_limit_final_usd']:,.2f}"
                 if usage["spend_limit_extensions"] else ""))
    if usage.get("spend_limit_stopped"):
        print("  Model switched off part way through the run at the spend limit;")
        print("  the remaining text was read by the local reader alone.")


def _share(part: int, whole: int) -> str:
    return f"{part:,} ({part / whole * 100:5.1f}%)" if whole else f"{part:,}"


def print_summary(manifest: Dict[str, Any], settings: Settings) -> None:
    statistics = manifest["statistics"]
    rows = statistics.get("rows_in", 0)

    print()
    print("=" * 79)
    print(f"{AGENT_NAME} - complete")
    print("=" * 79)
    print(f"  Run id               : {manifest['run_id']}")
    print(f"  Vocabulary version   : {manifest['lexicon_version']}")
    print(f"  Transaction rows in  : {rows:,}")
    print(f"  Rows written         : {statistics.get('rows_out', 0):,}")

    print()
    print("  Stage 1  invoice lines onto the transaction")
    print(f"    Matched on a line  : {_share(statistics.get('invoice_matched_line', 0), rows)}")
    print(f"    Matched on document: {_share(statistics.get('invoice_matched_header', 0), rows)}")
    print(f"    Not matched        : {_share(statistics.get('invoice_unmatched', 0), rows)}")

    print()
    print("  Stage 2  purchase order onto the widened table")
    print(f"    Matched on a line  : {_share(statistics.get('po_matched_line', 0), rows)}")
    print(f"    Matched on an order: {_share(statistics.get('po_matched_header', 0), rows)}")
    print(f"    Not matched        : {_share(statistics.get('po_unmatched', 0), rows)}")

    print()
    print("  Stage 3  free text read into columns")
    print(f"    Distinct strings   : {statistics.get('distinct_texts', 0):,}")
    print(f"    Below the floor    : {statistics.get('texts_below_floor', 0):,}")
    if statistics.get("texts_read_by_model"):
        print(f"    Read by the model  : {statistics['texts_read_by_model']:,}")

    stages = statistics.get("agent_stages") or {}
    if stages:
        print()
        print("  Stages 4 to 6  the agents onto the wide table")
        for label, stage in (("4  descriptions ", "max_stage4_enriched"),
                             ("5  purchase group", "max_stage5_grouped"),
                             ("6  standard item ", "max_stage6_standardised")):
            count = stages.get(f"{stage}_rows")
            if count is None:
                print(f"    {label}  : not written")
                continue
            print(f"    {label}  : {_share(stages.get(f'{stage}_annotated', 0), count)}")
        if stages.get("supplier_keys_resolved") is not None:
            print(f"    Supplier key       : "
                  f"{_share(stages['supplier_keys_resolved'], rows)}")
        for name, state in sorted((stages.get("agents") or {}).items()):
            if state != "ok":
                print(f"    {name}: {state}")

    missing = statistics.get("columns_missing") or []
    if missing:
        print()
        print(f"  Columns promised but missing ({len(missing)})")
        for name in missing[:12]:
            print(f"    {name}")
        if len(missing) > 12:
            print(f"    ... and {len(missing) - 12} more")

    carried = manifest.get("carried_over") or []
    if carried:
        print()
        print("  Carried over from an interrupted run")
        for name in carried:
            print(f"    {name}")
        print("    Rerun with --restart to rebuild these from stage 1.")

    if manifest.get("final_table"):
        print()
        print(f"  Table to use         : {manifest['final_table']}")
        print(f"  Columns              : {manifest.get('final_columns', 0):,}")

    advice = manifest.get("diagnostics", {}).get("advice") or []
    if advice:
        print()
        print("  Worth knowing")
        for note in advice:
            wrapped = re.findall(r".{1,70}(?:\s|$)", note)
            print(f"    - {wrapped[0].strip()}")
            for line in wrapped[1:]:
                if line.strip():
                    print(f"      {line.strip()}")

    print()
    print(f"  Output folder        : {settings.results_dir}")
    for name in manifest["outputs"]:
        print(f"    {name}")

    print_token_usage(statistics, settings)
    print()


def decide_resume(journal: RunJournal, settings: Settings) -> None:
    """Work out whether the previous run's finished stages can be carried over.

    Asked before anything is read, so that a run which is going to start again
    says so at the top rather than after the first stage has been repeated.
    """
    if not settings.resume:
        journal.discard()
        return

    obstacle = journal.load()
    if obstacle:
        if obstacle not in {"no previous run", "the previous run finished"}:
            LOGGER.info("Starting from stage 1: %s.", obstacle)
        journal.discard()
        return

    done = journal.done_stages()
    if not done:
        return

    finished = ", ".join(STAGE_LABELS.get(name, name) for name in done)
    stopped = journal.interrupted_at()

    print()
    LOGGER.warning("A previous run in this results folder did not finish.")
    LOGGER.warning("  Finished and reusable : %s", finished)
    if stopped:
        LOGGER.warning("  Stopped during        : %s",
                       STAGE_LABELS.get(stopped, stopped))

    if settings.interactive:
        if not ask_yes_no("Carry on from there rather than starting again", True):
            LOGGER.info("Starting again from stage 1.")
            journal.discard()
            return
    else:
        LOGGER.info("Carrying on from there. Use --restart to start again.")

    # The stage that was in flight when the run stopped left a part-written file
    # behind. It is not recorded as finished so it would never be reused, but it
    # is deleted anyway so that nobody opens it believing it to be a result.
    if stopped:
        for name in PARTIAL_OUTPUTS.get(stopped, ()):
            path = settings.results_dir / name
            if path.exists():
                try:
                    path.unlink()
                    LOGGER.info("Removed the part-written %s.", name)
                except OSError as error:
                    LOGGER.warning("Could not remove the part-written %s: %s",
                                   name, error)


def report_interruption(builder: Builder, journal: RunJournal,
                        settings: Settings) -> None:
    """Say what was kept and how to carry on, after Ctrl-C."""
    stage = journal.interrupted_at() or "the current stage"

    if builder.pipeline is not None:
        builder.pipeline.stop_running_agents()

    journal.stopped(journal.interrupted_at())

    done = journal.done_stages()
    print("\n")
    LOGGER.warning("Stopped during %s.", STAGE_LABELS.get(stage, stage))
    if done:
        LOGGER.warning("Kept, and carried over next time: %s.",
                       ", ".join(STAGE_LABELS.get(name, name) for name in done))
    else:
        LOGGER.warning("No stage had finished, so there is nothing to carry over.")

    if builder.pipeline is not None:
        finished = [run.name for run in builder.pipeline.runs if run.ok]
        if finished:
            LOGGER.warning("Agents already done: %s.", ", ".join(finished))

    LOGGER.warning("Run the same command again to carry on, or add --restart to "
                   "start from stage 1.")


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = build_parser().parse_args(argv)
    configure_logging(args.verbose)

    here = Path(__file__).resolve().parent
    env = {**load_dotenv(here / ".env"), **os.environ}

    try:
        settings = resolve_settings(args, env)
    except SystemExit as error:
        print(f"\n{error}\n")
        return 2

    LOGGER.info("Optional components: %s", ", ".join(
        f"{name}={'yes' if present else 'no'}"
        for name, present in sorted(describe_environment().items())))
    if settings.model.enabled:
        LOGGER.info("Language-model tier enabled: %s via %s",
                    settings.model.model, settings.model.backend)
        if settings.model.spend_limit:
            LOGGER.info("Spend alert set at $%.2f; the run will ask before going past it.",
                        settings.model.spend_limit)
        else:
            LOGGER.info("No spend alert set; the model tier will run unmetered.")

    # Claimed before the resume question, so a folder already in use is reported
    # at the top rather than after a stage has overwritten a shared file.
    lock = ResultsLock(settings.results_dir)
    refusal = lock.acquire()
    if refusal:
        LOGGER.error("Not starting: %s", refusal)
        return 1

    journal = RunJournal(settings)
    decide_resume(journal, settings)

    builder = Builder(settings)
    try:
        builder.load()
    except SystemExit as error:
        print(f"\n{error}\n")
        return 2

    journal.begin(builder.run_id or _now())
    try:
        builder.measure_keys()
        stage2_csv, _, stage2_columns = builder.wide_stage(journal)
        stage3_csv = builder.interpret_stage(journal, stage2_csv, stage2_columns)
        if settings.run_agents:
            journal.starting("agents")
            builder.enrich(stage3_csv)
            journal.completed("agents", [], {}, {})
        manifest = builder.write_manifest()
    except KeyboardInterrupt:
        report_interruption(builder, journal, settings)
        builder.close()
        lock.release()
        return 130
    finally:
        builder.close()
        lock.release()

    journal.complete()

    rows_in = builder.statistics.get("rows_in", 0)
    rows_out = builder.statistics.get("rows_out", 0)
    if rows_in != rows_out:
        # The whole design rests on this holding. If it ever does not, the
        # output is not a widened transaction table and must not be used as one.
        LOGGER.error("Row count changed during the build: %d in, %d out. "
                     "The output is not a faithful widening of the transactions.",
                     rows_in, rows_out)
        print_summary(manifest, settings)
        return 1

    # The same rule again, once per agent stage. A stage that lost or gained a
    # row is caught inside the merge, which then stops the chain; this reports it.
    stages = builder.statistics.get("agent_stages", {})
    ragged = sorted(name for name, value in stages.items()
                    if name.endswith("_rows") and value != rows_out)
    if ragged:
        LOGGER.error("An agent stage does not have %d rows: %s.",
                     rows_out, ", ".join(ragged))
        print_summary(manifest, settings)
        return 1

    print_summary(manifest, settings)
    if builder.statistics.get("columns_missing"):
        return 1
    if settings.run_agents and any(
            not run.ok for run in (builder.pipeline.runs if builder.pipeline else [])):
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
